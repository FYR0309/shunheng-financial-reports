# -*- coding: utf-8 -*-
"""Comprehensive self-check of generated statements"""
import openpyxl
from datetime import datetime
from collections import defaultdict

def sf(v):
    try: return float(v)
    except: return 0.0

def is_date(v):
    if v is None: return False
    if isinstance(v, datetime): return True
    s = str(v).strip()
    return len(s) >= 10 and s[4] == '-' and s[7] == '-'

print("=" * 60)
print("COMPREHENSIVE SELF-CHECK")
print("=" * 60)

# ==========================================
# 1. Read generated statements
# ==========================================
print("\n[1] Reading generated statements...")

wb_pl = openpyxl.load_workbook('顺恒2026年1-5月_利润表.xlsx', data_only=True)
ws_pl = wb_pl.active
pl_data = {}
for row in ws_pl.iter_rows(min_row=5, values_only=True):
    name = str(row[0]).strip() if row[0] else ''
    if name: pl_data[name] = {'cum': sf(row[2]), 'month': sf(row[3])}

wb_bs = openpyxl.load_workbook('顺恒2026年1-5月_资产负债表.xlsx', data_only=True)
ws_bs = wb_bs.active
bs_assets = {}; bs_liab = {}
for row in ws_bs.iter_rows(min_row=5, values_only=True):
    na = str(row[0]).strip() if row[0] else ''
    nl = str(row[4]).strip() if row[4] else ''
    if na: bs_assets[na] = {'end': sf(row[2]), 'begin': sf(row[3])}
    if nl: bs_liab[nl] = {'end': sf(row[6]), 'begin': sf(row[7])}

wb_cf = openpyxl.load_workbook('顺恒2026年1-5月_现金流量表.xlsx', data_only=True)
ws_cf = wb_cf.active
cf_data = {}
for row in ws_cf.iter_rows(min_row=5, values_only=True):
    name = str(row[0]).strip() if row[0] else ''
    if name: cf_data[name] = {'cum': sf(row[2]), 'month': sf(row[3])}

# Dec 2025 templates for format comparison (use xlrd for .xls)
import xlrd
wb_pl0 = xlrd.open_workbook('顺恒2025年12月_利润表.xls')
ws_pl0 = wb_pl0.sheet_by_index(0)
pl0_items = []
for r in range(5, ws_pl0.nrows):
    name = str(ws_pl0.cell_value(r, 0)).strip()
    if name: pl0_items.append(name)

wb_bs0 = xlrd.open_workbook('顺恒2025年12月_资产负债表（往来重分类）.xls')
ws_bs0 = wb_bs0.sheet_by_index(0)
bs0_assets = []; bs0_liab = []
for r in range(5, ws_bs0.nrows):
    na = str(ws_bs0.cell_value(r, 0)).strip()
    nl = str(ws_bs0.cell_value(r, 4)).strip()
    if na: bs0_assets.append(na)
    if nl: bs0_liab.append(nl)

# ==========================================
# 2. Cross-check against source data
# ==========================================
print("\n[2] Cross-check against source data...")
errors = []

# Sales
wb = openpyxl.load_workbook('顺恒2026年1-5月销售收入数据.xlsx', data_only=True)
ws = wb['信息汇总表']
src_sales = 0.0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not is_date(row[8]): continue
    src_sales += sf(row[16])

pl_rev = pl_data.get('一、营业收入', {}).get('cum', 0)
d = pl_rev - src_sales
print(f"  Revenue: PL={pl_rev:,.2f} vs Source={src_sales:,.2f} diff={d:,.2f}")
if abs(d) > 0.01: errors.append(f"Revenue mismatch: {d:,.2f}")

# COGS
wb = openpyxl.load_workbook('顺恒2026年1-5月成本费用数据.xlsx', data_only=True)
ws = wb['信息汇总表']
src_cogs = 0.0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not is_date(row[8]): continue
    code = str(row[9]) if row[9] else ''
    if code in ['1110701000000000000', '1110799000000000000']:
        src_cogs += sf(row[16])

pl_cogs = pl_data.get('减：营业成本', {}).get('cum', 0)
d = pl_cogs - src_cogs
print(f"  COGS: PL={pl_cogs:,.2f} vs Source={src_cogs:,.2f} diff={d:,.2f}")
if abs(d) > 0.01: errors.append(f"COGS mismatch: {d:,.2f}")

# Cash (use xlrd for .xls bank files)
import xlrd

def is_date_xl(v):
    if v is None or v == '': return False
    if isinstance(v, float) and v > 40000: return True
    s = str(v).strip()
    return len(s) >= 10 and s[4] == '-' and s[7] == '-'

nong_bal = 0; xin_bal = 0
wb = xlrd.open_workbook('顺恒2026年5月农行明细.xls')
ws = wb.sheet_by_index(0)
for r in range(3, ws.nrows):
    if not is_date_xl(ws.cell_value(r, 0)): continue
    nong_bal = sf(ws.cell_value(r, 3))
wb = xlrd.open_workbook('顺恒2026年5月农村信用社明细.xls')
ws = wb.sheet_by_index(0)
for r in range(4, ws.nrows):
    if not is_date_xl(ws.cell_value(r, 0)): continue
    xin_bal = sf(ws.cell_value(r, 3))
    break
bank_cash = nong_bal + xin_bal

bs_cash = bs_assets.get('货币资金', {}).get('end', 0)
d = bs_cash - bank_cash
print(f"  Cash: BS={bs_cash:,.2f} vs Bank={bank_cash:,.2f} diff={d:,.2f}")
if abs(d) > 0.01: errors.append(f"Cash mismatch: {d:,.2f}")

# Depreciation
depr_5m = 25886.46 * 5
acc_depr_beg = bs_assets.get('减：累计折旧', {}).get('begin', 0)
acc_depr_end = bs_assets.get('减：累计折旧', {}).get('end', 0)
d = acc_depr_end - acc_depr_beg - depr_5m
print(f"  Depreciation: change={acc_depr_end-acc_depr_beg:,.2f} vs expected={depr_5m:,.2f}")
if abs(d) > 0.01: errors.append(f"Depreciation mismatch: {d:,.2f}")

# Equity roll-forward
undist_beg = bs_liab.get('未分配利润', {}).get('begin', 0)
undist_end = bs_liab.get('未分配利润', {}).get('end', 0)
np_val = pl_data.get('四、净利润（净亏损以"-"号填列）', {}).get('cum', 0)
d = undist_end - undist_beg - np_val
print(f"  Equity: Undist change={undist_end-undist_beg:,.2f} vs Net Profit={np_val:,.2f} diff={d:,.2f}")
if abs(d) > 0.01: errors.append(f"Equity roll-forward mismatch: {d:,.2f}")

# ==========================================
# 3. Arithmetic checks
# ==========================================
print("\n[3] Arithmetic verification...")

rev = pl_data.get('一、营业收入', {}).get('cum', 0)
cogs = pl_data.get('减：营业成本', {}).get('cum', 0)
sur = pl_data.get('税金及附加', {}).get('cum', 0)
adm = pl_data.get('管理费用', {}).get('cum', 0)
fin = pl_data.get('财务费用', {}).get('cum', 0)
op = pl_data.get('二、营业利润（亏损以"-"号填列）', {}).get('cum', 0)
calc_op = rev - cogs - sur - adm - fin
print(f"  PL Operating Profit: stated={op:,.2f} calc={calc_op:,.2f} diff={op-calc_op:,.2f}")
if abs(op-calc_op) > 0.01: errors.append("PL operating profit arithmetic error")

np2 = pl_data.get('四、净利润（净亏损以"-"号填列）', {}).get('cum', 0)
print(f"  PL Net Profit: stated={np2:,.2f} calc same as op={calc_op:,.2f}")

ta = bs_assets.get('资产总计', {}).get('end', 0)
tl = bs_liab.get('负债合计', {}).get('end', 0)
te = bs_liab.get('所有者权益合计', {}).get('end', 0)
print(f"  BS: A={ta:,.2f} L={tl:,.2f} E={te:,.2f} A-L-E={ta-tl-te:,.2f}")
if abs(ta-tl-te) > 0.01: errors.append("BS not balanced")

ca = bs_assets.get('流动资产合计', {}).get('end', 0)
nca = bs_assets.get('非流动资产合计', {}).get('end', 0)
print(f"  BS: CA+NCA={ca+nca:,.2f} vs TA={ta:,.2f} diff={ca+nca-ta:,.2f}")
if abs(ca+nca-ta) > 0.01: errors.append("BS asset subtotals error")

cf_net = cf_data.get('四、现金净增加额', {}).get('cum', 0)
cf_beg = cf_data.get('加：期初现金余额', {}).get('cum', 0)
cf_end = cf_data.get('五、期末现金余额', {}).get('cum', 0)
print(f"  CF: Beg+Net={cf_beg+cf_net:,.2f} vs End={cf_end:,.2f} diff={cf_beg+cf_net-cf_end:,.2f}")
if abs(cf_beg+cf_net-cf_end) > 0.01: errors.append("CF ending balance error")

op_cf = cf_data.get('经营活动产生的现金流量净额', {}).get('cum', 0)
inv_cf = cf_data.get('投资活动产生的现金流量净额', {}).get('cum', 0)
fin_cf = cf_data.get('筹资活动产生的现金流量净额', {}).get('cum', 0)
print(f"  CF: Op+Inv+Fin={op_cf+inv_cf+fin_cf:,.2f} vs Net={cf_net:,.2f}")
if abs(op_cf+inv_cf+fin_cf-cf_net) > 0.01: errors.append("CF subtotal error")

# ==========================================
# 4. Format check
# ==========================================
print("\n[4] Format consistency...")

pl_items = list(pl_data.keys())
print(f"  PL items: 2025={len(pl0_items)} vs 2026={len(pl_items)}")
missing = set(pl0_items) - set(pl_items)
if missing:
    print(f"  PL items in 2025 but not 2026: {missing}")

bs_items_2026 = list(bs_assets.keys()) + list(bs_liab.keys())
bs_items_2025 = bs0_assets + bs0_liab
print(f"  BS items: 2025={len(bs_items_2025)} vs 2026={len(bs_items_2026)}")

# Key BS items must exist
must_have_bs = ['货币资金', '应收账款', '存货', '固定资产账面价值', '资产总计',
                '应付账款', '应交税费', '其他应付款', '负债合计',
                '实收资本（或股本）', '未分配利润', '所有者权益合计']
for item in must_have_bs:
    found = item in bs_assets or item in bs_liab
    if not found:
        errors.append(f"Missing BS item: {item}")

# ==========================================
# 5. Sanity checks
# ==========================================
print("\n[5] Sanity ratios...")

gm = (rev - cogs) / rev * 100 if rev else 0
print(f"  Gross Margin: {gm:.1f}%")

nm = np2 / rev * 100 if rev else 0
print(f"  Net Margin: {nm:.1f}%")

ar_ratio = adm / rev * 100 if rev else 0
print(f"  Admin/Revenue: {ar_ratio:.1f}%")

cl_liab = bs_liab.get('流动负债合计', {}).get('end', 0)
cr_val = ca / cl_liab if cl_liab else 0
print(f"  Current Ratio: {cr_val:.2f}")

dr = tl / ta * 100 if ta else 0
print(f"  Debt Ratio: {dr:.1f}%")

# Asset turnover
at = rev / ta if ta else 0
print(f"  Asset Turnover: {at:.2f}x")

print("\n" + "=" * 60)
if errors:
    print(f"FAILED! {len(errors)} errors found:")
    for e in errors:
        print(f"  [ERROR] {e}")
else:
    print("ALL CHECKS PASSED! No errors found.")
print("=" * 60)
