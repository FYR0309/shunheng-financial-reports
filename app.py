# -*- coding: utf-8 -*-
"""财务报表生成工具 — Streamlit Web Application.

Local web app for generating 小企业会计准则 financial statements.
All data stays on the user's computer. No network requests.
"""
import streamlit as st
import os
import json
import glob
import pandas as pd
from datetime import datetime
from generate_analysis_docx import generate_analysis_report

# ---- Page config ----
st.set_page_config(
    page_title='财务报表生成工具',
    page_icon='📊',
    layout='wide',
    initial_sidebar_state='expanded',
)

# ---- Paths ----
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
COMPANIES_DIR = os.path.join(DATA_DIR, 'companies')
UPLOADS_DIR = os.path.join(DATA_DIR, 'uploads')
OUTPUT_DIR = os.path.join(DATA_DIR, 'output')
LOGS_DIR = os.path.join(DATA_DIR, 'logs')

for d in [COMPANIES_DIR, UPLOADS_DIR, OUTPUT_DIR, LOGS_DIR]:
    os.makedirs(d, exist_ok=True)

# ---- Session state ----
DEFAULTS = {
    'current_company': None,
    'current_year': 2026,
    'current_month': 5,
    'page': '首页',
    'data_verified': False,
    'generated_pl_path': None,
    'generated_bs_path': None,
    'generated_cf_path': None,
    'generated_docx_path': None,
    'last_pl': None,
    'last_bs': None,
    'last_cf': None,
    'last_opening_bs': None,
    'last_period_label': None,
    'last_bs_date': None,
}
for key, val in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = key if key == 'page' else val

# ---- Helpers ----
def get_company_list():
    if not os.path.exists(COMPANIES_DIR):
        return []
    return sorted([d for d in os.listdir(COMPANIES_DIR)
                   if os.path.isdir(os.path.join(COMPANIES_DIR, d))])

def get_company_dir(name):
    return os.path.join(COMPANIES_DIR, name)

def load_company_config(name):
    path = os.path.join(get_company_dir(name), 'config.json')
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_company_config(name, config):
    d = get_company_dir(name)
    os.makedirs(d, exist_ok=True)
    with open(os.path.join(d, 'config.json'), 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def switch_page(name):
    st.session_state.page = name

# ---- Sidebar ----
with st.sidebar:
    st.title('📊 财务报表工具')

    companies = get_company_list()
    company_options = ['— 选择公司 —'] + companies
    current_idx = 0
    if st.session_state.current_company and st.session_state.current_company in companies:
        current_idx = companies.index(st.session_state.current_company) + 1

    selected_company = st.selectbox('当前公司', company_options, index=current_idx)
    if selected_company != '— 选择公司 —':
        st.session_state.current_company = selected_company

    c1, c2 = st.columns(2)
    with c1:
        years = list(range(2023, 2031))
        st.session_state.current_year = st.selectbox(
            '年份', years, index=years.index(st.session_state.current_year))
    with c2:
        st.session_state.current_month = st.selectbox(
            '月份', list(range(1, 13)), index=st.session_state.current_month - 1)

    st.divider()

    pages = ['首页', '公司档案', '科目映射', '数据导入', '数据核验', '报表生成', '分析导出', '历史对比']
    selected = st.radio('导航', pages, index=pages.index(st.session_state.page))
    st.session_state.page = selected

    st.divider()
    st.caption('所有数据仅在本地处理\n不上传任何服务器')

# ---- Import backend modules ----
from mapping_engine import MappingEngine
from data_extractor import InvoiceExtractor, BankExtractor, PayrollExtractor
from calc_engine import CalcEngine
from format_renderer import ReportRenderer
from validator import validate_import, validate_extraction, validate_balance, format_validation_message
from template_engine import (
    parse_custom_template,
    fill_custom_template,
    save_custom_config,
    load_custom_config,
)
import template_engine as te

# ================================================================
# PAGE: 首页
# ================================================================
if st.session_state.page == '首页':
    st.header('🏠 首页')

    if not st.session_state.current_company:
        st.info('👈 请先在左侧选择公司，或去「公司档案」创建新公司')
    else:
        company = st.session_state.current_company
        config = load_company_config(company)
        st.subheader(f'当前公司：{company}')
        if config.get('full_name'):
            st.caption(f"全称：{config['full_name']}")

        st.divider()

        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button('📤 导入数据', use_container_width=True, type='primary'):
                switch_page('数据导入')
                st.rerun()
        with col2:
            if st.button('📊 生成报表', use_container_width=True):
                if st.session_state.data_verified:
                    switch_page('报表生成')
                    st.rerun()
                else:
                    st.warning('请先导入并核验数据')
        with col3:
            if st.button('📝 导出报告', use_container_width=True):
                switch_page('分析导出')
                st.rerun()

        st.divider()
        st.subheader('历史报表')
        output_company_dir = os.path.join(OUTPUT_DIR, company)
        if os.path.exists(output_company_dir):
            past_files = sorted(
                [f for f in os.listdir(output_company_dir) if f.endswith('.xlsx')],
                reverse=True)[:12]
            if past_files:
                for f in past_files:
                    st.write(f'✅ {f}')
            else:
                st.caption('暂无历史报表')
        else:
            st.caption('暂无历史报表')

# ================================================================
# PAGE: 公司档案
# ================================================================
elif st.session_state.page == '公司档案':
    st.header('📋 公司档案')

    tab1, tab2 = st.tabs(['公司信息', '年初余额'])

    with tab1:
        action = st.radio('操作', ['选择已有公司', '创建新公司'], horizontal=True)

        if action == '创建新公司':
            new_name = st.text_input('公司简称（用于文件夹命名）', placeholder='例如：顺恒废旧')
            full_name = st.text_input('公司全称（将显示在报表上）',
                                      placeholder='例如：来宾市顺恒废旧汽车回收有限公司')
            tax_id = st.text_input('纳税人识别号', placeholder='可选')

            if st.button('创建公司', type='primary'):
                if new_name:
                    config = {
                        'full_name': full_name,
                        'tax_id': tax_id,
                        'created_at': datetime.now().isoformat(),
                        'template_type': 'preset',
                        'opening_bs': {},
                    }
                    save_company_config(new_name, config)
                    st.session_state.current_company = new_name
                    st.success(f'公司「{new_name}」创建成功！')
                    st.rerun()
                else:
                    st.error('请输入公司简称')
        else:
            companies = get_company_list()
            if companies:
                idx = companies.index(st.session_state.current_company) if \
                    st.session_state.current_company in companies else 0
                sel = st.selectbox('选择公司', companies, index=idx)
                if sel:
                    st.session_state.current_company = sel
                    config = load_company_config(sel)
                    st.write(f"全称：{config.get('full_name', '未设置')}")
            else:
                st.info('暂无公司档案，请先创建')

    with tab2:
        if not st.session_state.current_company:
            st.info('请先选择或创建公司')
        else:
            company = st.session_state.current_company
            config = load_company_config(company)
            opening = config.get('opening_bs', {})

            st.subheader('年初资产负债表余额')
            st.caption('填写去年12月31日的资产负债表期末数。填一次即可。')

            col1, col2 = st.columns(2)
            with col1:
                st.markdown('**资产**')
                opening['货币资金'] = st.number_input('货币资金', value=float(opening.get('货币资金', 0)), step=1000.0, format='%.2f', key='bs_cash')
                opening['应收账款'] = st.number_input('应收账款', value=float(opening.get('应收账款', 0)), step=1000.0, format='%.2f', key='bs_ar')
                opening['预付账款'] = st.number_input('预付账款', value=float(opening.get('预付账款', 0)), step=1000.0, format='%.2f', key='bs_prepay')
                opening['其他应收款'] = st.number_input('其他应收款', value=float(opening.get('其他应收款', 0)), step=1000.0, format='%.2f', key='bs_orecv')
                opening['存货'] = st.number_input('存货', value=float(opening.get('存货', 0)), step=1000.0, format='%.2f', key='bs_inv')
                opening['固定资产原价'] = st.number_input('固定资产原价', value=float(opening.get('固定资产原价', 0)), step=10000.0, format='%.2f', key='bs_fa')
                opening['减：累计折旧'] = st.number_input('减：累计折旧', value=float(opening.get('减：累计折旧', 0)), step=1000.0, format='%.2f', key='bs_depr')
                opening['长期待摊费用'] = st.number_input('长期待摊费用', value=float(opening.get('长期待摊费用', 0)), step=1000.0, format='%.2f', key='bs_ltd')

            with col2:
                st.markdown('**负债及所有者权益**')
                opening['应付账款'] = st.number_input('应付账款', value=float(opening.get('应付账款', 0)), step=1000.0, format='%.2f', key='bs_ap')
                opening['预收账款'] = st.number_input('预收账款', value=float(opening.get('预收账款', 0)), step=1000.0, format='%.2f', key='bs_pr')
                opening['应付职工薪酬'] = st.number_input('应付职工薪酬', value=float(opening.get('应付职工薪酬', 0)), step=1000.0, format='%.2f', key='bs_payable')
                opening['应交税费'] = st.number_input('应交税费', value=float(opening.get('应交税费', 0)), step=1000.0, format='%.2f', key='bs_tax')
                opening['其他应付款'] = st.number_input('其他应付款', value=float(opening.get('其他应付款', 0)), step=10000.0, format='%.2f', key='bs_other')
                opening['实收资本（或股本）'] = st.number_input('实收资本（或股本）', value=float(opening.get('实收资本（或股本）', 550000)), step=10000.0, format='%.2f', key='bs_capital')
                opening['未分配利润'] = st.number_input('未分配利润', value=float(opening.get('未分配利润', 0)), step=10000.0, format='%.2f', key='bs_undist')

            if st.button('保存年初余额', type='primary'):
                config['opening_bs'] = opening
                save_company_config(company, config)
                st.success('年初余额已保存！')

# ================================================================
# PAGE: 科目映射
# ================================================================
elif st.session_state.page == '科目映射':
    st.header('⚙️ 科目映射')

    if not st.session_state.current_company:
        st.info('请先选择公司')
    else:
        company = st.session_state.current_company
        company_dir = get_company_dir(company)
        engine = MappingEngine(company_dir)

        statement = st.selectbox('选择报表', ['利润表 (pl)', '资产负债表 (bs)'])
        st_type = 'pl' if '利润' in statement else 'bs'

        # Show current mappings
        current = engine.get_mappings(st_type)
        if current:
            st.subheader('当前映射')
            for m in current:
                codes = ', '.join(m.get('account_codes', []))
                rtype = m['rule_type']
                st.write(f"- **{m['report_line']}** ← `{codes}` ({rtype})")

        st.divider()

        # Add manual mapping
        st.subheader('手动添加映射')
        cfg = te.load_preset(st_type)
        if st_type == 'bs':
            item_names = sorted(set(
                [it['left'] for it in cfg['items'] if it['left']] +
                [it['right'] for it in cfg['items'] if it['right']]
            ))
        else:
            item_names = sorted(set(it['name'] for it in cfg['items'] if it['name']))

        col1, col2 = st.columns(2)
        with col1:
            target_line = st.selectbox('报表行项目', ['— 选择 —'] + item_names, key='map_target')
        with col2:
            new_codes = st.text_input('科目编码（多个用逗号分隔）', placeholder='例如: 1110701000000000000, 1110799...')

        if st.button('添加到映射', type='primary'):
            if target_line != '— 选择 —' and new_codes.strip():
                codes = [c.strip() for c in new_codes.split(',') if c.strip()]
                engine.add_code_mapping(st_type, target_line, codes)
                st.success(f'已添加：{target_line} ← {codes}')
                st.rerun()
            else:
                st.error('请选择报表行并输入科目编码')

        # Discover codes from uploaded sample
        st.divider()
        st.subheader('从样本文件发现科目编码')
        st.caption('上传一份发票文件，系统会扫描所有科目编码并显示未配置的。')
        sample_file = st.file_uploader('上传样本文件', type=['xlsx', 'xls'], key='mapping_sample')

        if sample_file:
            tmp_path = os.path.join(UPLOADS_DIR, f'_sample_{company}.xlsx')
            with open(tmp_path, 'wb') as f:
                f.write(sample_file.read())

            try:
                ie = InvoiceExtractor(tmp_path)
                by_code = ie.get_total_by_code()
                st.success(f'发现 {len(by_code)} 个科目编码')

                unmapped = engine.find_unmapped_codes(st_type, list(by_code.keys()))
                if unmapped:
                    st.warning(f'{len(unmapped)} 个编码未配置映射：')
                    for code in unmapped:
                        col1, col2 = st.columns([1, 2])
                        with col1:
                            st.code(code)
                        with col2:
                            quick_target = st.selectbox(
                                f'映射到 →', ['— 选择 —'] + item_names,
                                key=f'quick_{code}')
                            if quick_target != '— 选择 —':
                                if st.button(f'确认映射', key=f'btn_{code}'):
                                    engine.add_code_mapping(st_type, quick_target, [code])
                                    st.rerun()
                else:
                    st.success('所有编码已配置映射！✅')
            except Exception as e:
                st.error(f'读取文件失败：{e}')
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

# ================================================================
# PAGE: 数据导入
# ================================================================
elif st.session_state.page == '数据导入':
    st.header('📤 数据导入')

    if not st.session_state.current_company:
        st.info('请先选择公司')
    else:
        company = st.session_state.current_company
        year = st.session_state.current_year
        month = st.session_state.current_month

        st.subheader(f'导入 {year}年{month}月 数据')
        st.caption('上传以下文件。支持 .xls 和 .xlsx 格式。')

        col1, col2 = st.columns(2)
        with col1:
            sales_file = st.file_uploader('📄 销售收入发票', type=['xlsx', 'xls'], key='sales')
            costs_file = st.file_uploader('📄 成本费用发票', type=['xlsx', 'xls'], key='costs')
        with col2:
            bank_nong = st.file_uploader('🏦 农行流水 (.xls)', type=['xls'], key='nong')
            bank_xin = st.file_uploader('🏦 信用社流水 (.xls)', type=['xls'], key='xin')
            payroll_file = st.file_uploader('👤 工资薪金表', type=['xlsx', 'xls'], key='payroll')

        if st.button('保存文件并进入核验 →', type='primary'):
            if not any([sales_file, costs_file, bank_nong, bank_xin, payroll_file]):
                st.error('请至少上传一个文件')
            else:
                saved = []
                upload_dir = os.path.join(UPLOADS_DIR, company, f'{year}-{month:02d}')
                os.makedirs(upload_dir, exist_ok=True)

                for label, file_obj in [
                    ('sales', sales_file), ('costs', costs_file),
                    ('nong', bank_nong), ('xin', bank_xin), ('payroll', payroll_file)
                ]:
                    if file_obj:
                        ext = file_obj.name.split('.')[-1]
                        fname = f'{label}.{ext}'
                        with open(os.path.join(upload_dir, fname), 'wb') as f:
                            f.write(file_obj.read())
                        saved.append(fname)

                st.success(f'已保存 {len(saved)} 个文件：{", ".join(saved)}')
                st.session_state.data_verified = False
                switch_page('数据核验')
                st.rerun()

# ================================================================
# PAGE: 数据核验
# ================================================================
elif st.session_state.page == '数据核验':
    st.header('✅ 数据核验')

    if not st.session_state.current_company:
        st.info('请先选择公司')
    else:
        company = st.session_state.current_company
        year = st.session_state.current_year
        month = st.session_state.current_month
        upload_dir = os.path.join(UPLOADS_DIR, company, f'{year}-{month:02d}')

        if not os.path.exists(upload_dir) or not os.listdir(upload_dir):
            st.warning('未找到本月上传的数据。请先到「数据导入」上传文件。')
            if st.button('去导入数据'):
                switch_page('数据导入')
                st.rerun()
        else:
            st.subheader('📊 数据概览')

            # Extract and display
            extraction = {'sales': {}, 'costs': {}, 'bank': {}}
            bank_end = 0.0

            # Sales
            sales_path = os.path.join(upload_dir, 'sales.xlsx')
            if os.path.exists(sales_path):
                ie = InvoiceExtractor(sales_path)
                data, skipped = ie.extract()
                extraction['sales'] = data
                total_sales = sum(sum(r['amount'] for r in recs) for recs in data.values())
                count = sum(len(recs) for recs in data.values())
                st.metric('销售收入', f'¥{total_sales:,.2f}', f'{count}张发票')

            # Costs
            costs_path = os.path.join(upload_dir, 'costs.xlsx')
            if os.path.exists(costs_path):
                ie = InvoiceExtractor(costs_path)
                data, skipped = ie.extract()
                extraction['costs'] = data
                total_costs = sum(sum(r['amount'] for r in recs) for recs in data.values())
                count = sum(len(recs) for recs in data.values())
                st.metric('成本费用', f'¥{total_costs:,.2f}', f'{count}张发票')

            # Bank
            c1, c2 = st.columns(2)
            nong_result = None
            with c1:
                nong_path = os.path.join(upload_dir, 'nong.xls')
                if os.path.exists(nong_path):
                    be = BankExtractor(nong_path, BankExtractor.BANK_NONGHANG)
                    nong_result = be.extract()
                    extraction['bank']['农行'] = nong_result
                    bank_end += nong_result['end_balance']
                    st.metric('农行余额', f'¥{nong_result["end_balance"]:,.2f}')
            with c2:
                xin_path = os.path.join(upload_dir, 'xin.xls')
                if os.path.exists(xin_path):
                    be = BankExtractor(xin_path, BankExtractor.BANK_XINYONGSHE)
                    xin_result = be.extract()
                    extraction['bank']['信用社'] = xin_result
                    bank_end += xin_result['end_balance']
                    st.metric('信用社余额', f'¥{xin_result["end_balance"]:,.2f}')
            st.metric('🏦 银行余额合计', f'¥{bank_end:,.2f}')

            # Payroll
            payroll_path = os.path.join(upload_dir, 'payroll.xlsx')
            if os.path.exists(payroll_path):
                pe = PayrollExtractor(payroll_path)
                pr = pe.extract()
                extraction['payroll'] = pr
                st.metric('工资应发', f'¥{pr["gross_pay"]:,.2f}')

            # Validation
            st.divider()
            st.subheader('🔍 核验结果')

            company_dir = get_company_dir(company)
            engine = MappingEngine(company_dir)
            passed, errors, warnings = validate_extraction(extraction, engine, 'pl')

            if not errors and not warnings:
                st.success('✅ 核验通过，可以生成报表')
                st.session_state.data_verified = True
            else:
                if not errors:
                    st.success('✅ 未发现严重问题')
                    st.session_state.data_verified = True
                else:
                    st.error(f'🔴 发现 {len(errors)} 个问题：')
                    for e in errors:
                        st.error(format_validation_message(e))

                if warnings:
                    for w in warnings:
                        st.warning(format_validation_message(w))

            c1, c2 = st.columns(2)
            with c1:
                if st.button('← 返回修改数据', use_container_width=True):
                    switch_page('数据导入')
                    st.rerun()
            with c2:
                if st.button('确认无误，生成报表 →', type='primary', use_container_width=True,
                            disabled=not st.session_state.data_verified):
                    switch_page('报表生成')
                    st.rerun()

# ================================================================
# PAGE: 报表生成
# ================================================================
elif st.session_state.page == '报表生成':
    st.header('📊 报表生成')

    if not st.session_state.current_company:
        st.info('请先选择公司')
    elif not st.session_state.data_verified:
        st.warning('请先完成数据核验')
        if st.button('去核验数据'):
            switch_page('数据核验')
            st.rerun()
    else:
        company = st.session_state.current_company
        year = st.session_state.current_year
        month = st.session_state.current_month
        company_dir = get_company_dir(company)
        upload_dir = os.path.join(UPLOADS_DIR, company, f'{year}-{month:02d}')

        # ---- Custom Template Upload ----
        with st.expander('📋 自定义模板（可选）', expanded=False):
            st.caption('上传你自己的 xlsx 报表模板，系统会自动匹配数据字段并填入。')
            c_t1, c_t2, c_t3 = st.columns(3)
            stypes = ['pl', 'bs', 'cf']
            stype_labels = {'pl': '利润表', 'bs': '资产负债表', 'cf': '现金流量表'}

            for idx, (col, sty) in enumerate(zip([c_t1, c_t2, c_t3], stypes)):
                with col:
                    uploaded = st.file_uploader(
                        f'{stype_labels[sty]}模板', type=['xlsx', 'xls'],
                        key=f'custom_tpl_{sty}',
                        help=f'上传自定义{stype_labels[sty]}格式'
                    )
                    if uploaded:
                        tpl_dir = os.path.join(DATA_DIR, 'templates', company)
                        os.makedirs(tpl_dir, exist_ok=True)
                        tpl_path = os.path.join(tpl_dir, f'{sty}_custom.xlsx')
                        with open(tpl_path, 'wb') as f:
                            f.write(uploaded.getbuffer())

                        try:
                            parsed = parse_custom_template(tpl_path, statement_type=sty)
                            save_custom_config(company, sty, parsed['field_map'],
                                               uploaded.name)
                            matched = len(parsed['field_map'])
                            total = matched + len(parsed['unmatched'])
                            st.success(f'✅ 已解析: {matched}/{total} 行匹配 ({stype_labels[sty]})')
                            if parsed['unmatched']:
                                st.caption(f'未匹配: {len(parsed["unmatched"])} 行')
                        except Exception as e:
                            st.error(f'解析失败: {e}')

            # Show saved custom templates
            saved = []
            for sty in stypes:
                if load_custom_config(company, sty):
                    saved.append(stype_labels[sty])
            if saved:
                st.info(f'已保存自定义模板: {", ".join(saved)}')

        if st.button('🚀 生成三大报表', type='primary', use_container_width=True):
            with st.spinner('正在生成报表...'):
                config = load_company_config(company)
                opening_bs = config.get('opening_bs', {})
                if not opening_bs:
                    st.error('请先到「公司档案」页填写年初余额')
                    st.stop()

                # Extract
                results = {}
                sales_path = os.path.join(upload_dir, 'sales.xlsx')
                costs_path = os.path.join(upload_dir, 'costs.xlsx')
                nong_path = os.path.join(upload_dir, 'nong.xls')
                xin_path = os.path.join(upload_dir, 'xin.xls')
                payroll_path = os.path.join(upload_dir, 'payroll.xlsx')

                if os.path.exists(sales_path):
                    results['sales'] = InvoiceExtractor(sales_path).extract()[0]
                if os.path.exists(costs_path):
                    results['costs'] = InvoiceExtractor(costs_path).extract()[0]

                bank_data = {}
                bank_end = 0.0
                for label, btype, fpath in [
                    ('农行', BankExtractor.BANK_NONGHANG, nong_path),
                    ('信用社', BankExtractor.BANK_XINYONGSHE, xin_path),
                ]:
                    if os.path.exists(fpath):
                        be = BankExtractor(fpath, btype)
                        r = be.extract()
                        bank_data[label] = r
                        bank_end += r['end_balance']

                payroll_data = {}
                if os.path.exists(payroll_path):
                    payroll_data[month] = PayrollExtractor(payroll_path).extract()

                # Calculate
                engine = MappingEngine(company_dir)
                calc = CalcEngine(opening_bs)
                num_months = month

                pl = calc.calculate_pl(
                    {'sales': results.get('sales', {}), 'costs': results.get('costs', {})},
                    payroll_data,
                    engine.get_mappings('pl'),
                    {month: bank_data},
                    num_months,
                )

                fa_add = pl['_detail']['fa_add']
                bs = calc.calculate_bs(pl, bank_end, fa_add, num_months)

                # Validate
                v_passed, v_errors, v_warnings = validate_balance(pl, bs)

                # Show metrics
                st.success('报表计算完成！')
                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    st.metric('营业收入', f'¥{pl["revenue"]:,.2f}' if pl['revenue'] else '¥0.00')
                with c2:
                    st.metric('净利润', f'¥{pl["net_profit"]:,.2f}' if pl['net_profit'] else '¥0.00')
                with c3:
                    st.metric('资产总计', f'¥{bs["total_assets"]:,.2f}' if bs['total_assets'] else '¥0.00')
                with c4:
                    st.metric('银行余额', f'¥{bank_end:,.2f}')

                # Validation
                if v_passed:
                    st.success(f'✅ 校验通过 — 资产负债表平衡')
                if v_errors:
                    for e in v_errors:
                        st.error(format_validation_message(e))
                if v_warnings:
                    for w in v_warnings:
                        st.warning(format_validation_message(w))

                # ---- Render helpers ----
                period_label = f'{year}年1-{month}月'
                bs_date = f'{year}年{month}月28日'
                output_dir = os.path.join(OUTPUT_DIR, company)
                os.makedirs(output_dir, exist_ok=True)
                full_name = config.get('full_name', company)
                tpl_base = os.path.join(DATA_DIR, 'templates', company)

                def _render_or_fill(stype, preset_map, stype_label):
                    """Render using custom template if available, else preset."""
                    custom_cfg = load_custom_config(company, stype)
                    tpl_path = os.path.join(tpl_base, f'{stype}_custom.xlsx') if tpl_base else None
                    output_path = os.path.join(output_dir,
                        f'{company}_{year}年1-{month}月_{stype_label}.xlsx')

                    if custom_cfg and tpl_path and os.path.exists(tpl_path):
                        # Use custom template
                        fill_custom_template(
                            tpl_path, custom_cfg['field_map'], preset_map,
                            output_path, company_name=full_name, period_label=period_label)
                        return output_path
                    else:
                        # Use preset renderer
                        cfg = te.load_preset(stype)
                        renderer = ReportRenderer(cfg)
                        renderer.render(full_name, period_label, preset_map, output_path)
                        return output_path

                # PL data map
                pl_map = {
                    '一、营业收入': {0: pl['revenue'], 1: pl['revenue']},
                    '减：营业成本': {0: pl['cogs'], 1: pl['cogs']},
                    '税金及附加': {0: pl['surcharges'], 1: pl['surcharges']},
                    '管理费用': {0: pl['admin_exp'], 1: pl['admin_exp']},
                    '财务费用': {0: pl['fin_exp'], 1: pl['fin_exp']},
                    '二、营业利润（亏损以"-"号填列）': {0: pl['oper_profit'], 1: pl['oper_profit']},
                    '三、利润总额（亏损总额以"-"号填列）': {0: pl['total_profit'], 1: pl['total_profit']},
                    '四、净利润（净亏损以"-"号填列）': {0: pl['net_profit'], 1: pl['net_profit']},
                }
                pl_path = _render_or_fill('pl', pl_map, '利润表')
                st.session_state.generated_pl_path = pl_path

                # BS data map
                bs_map = {}
                asset_items = [('货币资金', bs.get('cash'), opening_bs.get('货币资金')),
                    ('应收账款', bs.get('ar'), opening_bs.get('应收账款')),
                    ('预付账款', bs.get('prepay'), opening_bs.get('预付账款')),
                    ('其他应收款', bs.get('other_recv'), opening_bs.get('其他应收款')),
                    ('存货', bs.get('inventory'), opening_bs.get('存货')),
                    ('流动资产合计', bs.get('curr_assets'), opening_bs.get('流动资产合计')),
                    ('固定资产原价', bs.get('fa_orig'), opening_bs.get('固定资产原价')),
                    ('减：累计折旧', bs.get('acc_depr'), bs.get('acc_depr_beg')),
                    ('固定资产账面价值', bs.get('fa_net'), opening_bs.get('固定资产账面价值')),
                    ('长期待摊费用', bs.get('ltd'), opening_bs.get('长期待摊费用')),
                    ('非流动资产合计', bs.get('non_curr_assets'), opening_bs.get('非流动资产合计')),
                    ('资产总计', bs.get('total_assets'), opening_bs.get('资产总计')),
                ]
                liab_items = [('应付账款', bs.get('ap'), opening_bs.get('应付账款')),
                    ('预收账款', bs.get('pr'), opening_bs.get('预收账款')),
                    ('应付职工薪酬', bs.get('payroll_payable'), opening_bs.get('应付职工薪酬')),
                    ('应交税费', bs.get('tax_payable'), opening_bs.get('应交税费')),
                    ('其他应付款', bs.get('other_pay'), opening_bs.get('其他应付款')),
                    ('流动负债合计', bs.get('curr_liab'), opening_bs.get('流动负债合计')),
                    ('负债合计', bs.get('total_liab'), opening_bs.get('负债合计')),
                    ('实收资本（或股本）', bs.get('capital'), opening_bs.get('实收资本（或股本）')),
                    ('未分配利润', bs.get('undist_profit_end'), bs.get('undist_profit_beg')),
                    ('所有者权益（或股东权益）合计', bs.get('total_equity'), opening_bs.get('所有者权益（或股东权益）合计')),
                    ('负债和所有者权益（或股东权益）总计', bs.get('total_le'), opening_bs.get('负债和所有者权益（或股东权益）总计')),
                ]
                for name, end, beg in asset_items + liab_items:
                    bs_map[name] = {0: end, 1: beg}

                bs_path = _render_or_fill('bs', bs_map, '资产负债表')
                st.session_state.generated_bs_path = bs_path

                # CF: compute then render
                cf = calc.calculate_cf(pl, bs, {month: bank_data},
                                       {'sales': results.get('sales', {}), 'costs': results.get('costs', {})},
                                       payroll_data, fa_add, num_months)

                # Build CF data map: key = template item name, value = {0: cumulative, 1: monthly}
                cf_items = [
                    '销售产成品、商品、提供劳务收到的现金',
                    '收到的其他与经营活动有关的现金',
                    '购买原材料、商品、接受劳务支付的现金',
                    '支付的职工薪酬',
                    '支付的税费',
                    '支付的其他与经营活动有关的现金',
                    '经营活动产生的现金流量净额',
                    '收回短期投资、长期债券投资和长期股权投资收到的现金',
                    '取得投资收益收到的现金',
                    '处置固定资产、无形资产和其他非流动资产收回的现金净额',
                    '短期投资、长期债券投资和长期股权投资支付的现金',
                    '购建固定资产、无形资产和其他非流动资产支付的现金',
                    '投资活动产生的现金流量净额',
                    '取得借款收到的现金',
                    '吸收投资者投资收到的现金',
                    '偿还借款本金支付的现金',
                    '偿还借款利息支付的现金',
                    '分配利润支付的现金',
                    '筹资活动产生的现金流量净额',
                    '四、现金净增加额',
                    '加：期初现金余额',
                    '五、期末现金余额',
                    # Indirect method (supplementary)
                    '净利润',
                    '加：计提的资产减值准备',
                    '固定资产折旧',
                    '无形资产摊销',
                    '长期待摊费用摊销',
                    '待摊费用减少（减：增加）',
                    '预提费用增加（减：减少）',
                    '处置固定资产、无形资产和其他非流动资产损失（减：收益）',
                    '固定资产报废损失',
                    '财务费用',
                    '投资损失（减：收益）',
                    '递延税款贷项（减：借项）',
                    '存货的减少（减：增加）',
                    '经营性应收项目的减少（减：增加）',
                    '经营性应付项目的增加（减：减少）',
                    '其他',
                    '债务转为资本',
                    '一年内到期的可转换公司债券',
                    '融资租入固定资产',
                    '现金的期末余额',
                    '减：现金的期初余额',
                    '加：现金等价物的期末余额',
                    '减：现金等价物的期初余额',
                    '现金及现金等价物净增加额',
                ]
                cf_map = {}
                for key in cf_items:
                    val = cf.get(key)
                    if val is not None:
                        cf_map[key] = {0: val, 1: round(val / num_months, 2) if num_months else None}

                cf_path = _render_or_fill('cf', cf_map, '现金流量表')
                st.session_state.generated_cf_path = cf_path

                # Save results for analysis report
                st.session_state.last_pl = pl
                st.session_state.last_bs = bs
                st.session_state.last_cf = cf
                st.session_state.last_opening_bs = opening_bs
                st.session_state.last_period_label = period_label
                st.session_state.last_bs_date = bs_date

                # Save summary JSON for historical comparison
                import datetime as _dt
                summary = {
                    'generated_at': _dt.datetime.now().isoformat(),
                    'period': period_label,
                    'year': year, 'month': month,
                    'pl': {k: v for k, v in pl.items() if not k.startswith('_')},
                    'bs': {k: v for k, v in bs.items() if not k.startswith('_')},
                    'cf': {k: v for k, v in cf.items() if not k.startswith('_')},
                }
                summary_path = os.path.join(output_dir, f'summary_{year}_{month:02d}.json')
                with open(summary_path, 'w', encoding='utf-8') as sf:
                    json.dump(summary, sf, ensure_ascii=False, indent=2)

                st.success(f'✅ 报表已生成！')
                st.info(f'利润表: {pl_path}\n资产负债表: {bs_path}')

                # Preview PL
                st.divider()
                st.subheader('利润表预览')
                df_pl = pd.read_excel(pl_path, header=None)
                ncols_pl = min(4, df_pl.shape[1])
                st.dataframe(df_pl.iloc[:25, :ncols_pl], use_container_width=True, height=400)

# ================================================================
# PAGE: 分析导出
# ================================================================
elif st.session_state.page == '分析导出':
    st.header('📝 分析导出')

    if not st.session_state.current_company:
        st.info('请先选择公司')
    else:
        company = st.session_state.current_company

        st.subheader('下载报表文件')
        c1, c2, c3 = st.columns(3)

        with c1:
            if st.session_state.generated_pl_path and os.path.exists(st.session_state.generated_pl_path):
                with open(st.session_state.generated_pl_path, 'rb') as f:
                    st.download_button('📥 下载利润表', f.read(),
                        os.path.basename(st.session_state.generated_pl_path),
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True)
            else:
                st.button('📥 利润表（未生成）', disabled=True, use_container_width=True)

        with c2:
            if st.session_state.generated_bs_path and os.path.exists(st.session_state.generated_bs_path):
                with open(st.session_state.generated_bs_path, 'rb') as f:
                    st.download_button('📥 下载资产负债表', f.read(),
                        os.path.basename(st.session_state.generated_bs_path),
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True)
            else:
                st.button('📥 资产负债表（未生成）', disabled=True, use_container_width=True)

        with c3:
            if st.session_state.generated_cf_path and os.path.exists(st.session_state.generated_cf_path):
                with open(st.session_state.generated_cf_path, 'rb') as f:
                    st.download_button('📥 下载现金流量表', f.read(),
                        os.path.basename(st.session_state.generated_cf_path),
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True)
            else:
                st.button('📥 现金流量表（未生成）', disabled=True, use_container_width=True)

        st.divider()
        st.subheader('📝 分析报告')

        col_r1, col_r2 = st.columns([1, 2])
        with col_r1:
            if st.session_state.last_pl and st.session_state.last_bs and st.session_state.last_cf:
                if st.button('📄 生成分析报告（Word）', type='primary', use_container_width=True):
                    company = st.session_state.current_company
                    output_dir = os.path.join(OUTPUT_DIR, company)
                    os.makedirs(output_dir, exist_ok=True)
                    full_name = config.get('full_name', company)
                    period_label = st.session_state.last_period_label
                    bs_date = st.session_state.last_bs_date

                    docx_path = os.path.join(
                        output_dir, f'{company}_{period_label}_财务分析报告.docx')

                    with st.spinner('正在生成分析报告...'):
                        generate_analysis_report(
                            full_name, period_label, bs_date,
                            st.session_state.last_pl,
                            st.session_state.last_bs,
                            st.session_state.last_cf,
                            st.session_state.last_opening_bs,
                            docx_path)
                    st.session_state.generated_docx_path = docx_path
                    st.success('✅ 分析报告已生成！')

                if st.session_state.generated_docx_path and \
                   os.path.exists(st.session_state.generated_docx_path):
                    with open(st.session_state.generated_docx_path, 'rb') as f:
                        st.download_button(
                            '📥 下载分析报告', f.read(),
                            os.path.basename(st.session_state.generated_docx_path),
                            mime='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                            use_container_width=True)
            else:
                st.button('📄 生成分析报告（需先生成报表）', disabled=True, use_container_width=True)
                st.caption('请先在「报表生成」页面生成三大报表。')

        if st.session_state.generated_pl_path:
            st.info('✅ 报表已生成，可使用上方按钮下载。')

    st.divider()
    out_dir = os.path.join(OUTPUT_DIR, st.session_state.current_company) if st.session_state.current_company else OUTPUT_DIR
    st.caption(f'报表保存位置：{out_dir}')

# ================================================================
# PAGE: 历史对比
# ================================================================
elif st.session_state.page == '历史对比':
    st.header('📈 历史对比')

    if not st.session_state.current_company:
        st.info('请先选择公司')
    else:
        company = st.session_state.current_company
        output_dir = os.path.join(OUTPUT_DIR, company)

        # Scan for available summaries
        summaries = {}
        if os.path.exists(output_dir):
            for fname in sorted(os.listdir(output_dir)):
                if fname.startswith('summary_') and fname.endswith('.json'):
                    path = os.path.join(output_dir, fname)
                    try:
                        with open(path, 'r', encoding='utf-8') as sf:
                            s = json.load(sf)
                        period = s.get('period', fname)
                        summaries[period] = s
                    except Exception:
                        pass

        if len(summaries) < 2:
            st.info(f'需要至少生成2个月份的报表才能对比。当前: {len(summaries)} 个')
            if summaries:
                st.caption(f'可用月份: {", ".join(summaries.keys())}')
        else:
            periods = list(summaries.keys())
            selected = st.multiselect(
                '选择对比月份（可多选）', periods,
                default=periods[-min(6, len(periods)):],
                help='选择要对比的月份，按选择顺序排列'
            )

            if len(selected) < 2:
                st.info('请至少选择2个月份')
            else:
                # ---- PL Comparison ----
                st.subheader('📊 利润表对比')

                pl_fields = [
                    ('revenue', '营业收入'),
                    ('cogs', '营业成本'),
                    ('surcharges', '税金及附加'),
                    ('selling_exp', '销售费用'),
                    ('admin_exp', '管理费用'),
                    ('fin_exp', '财务费用'),
                    ('oper_profit', '营业利润'),
                    ('total_profit', '利润总额'),
                    ('net_profit', '净利润'),
                ]

                # Build comparison table
                pl_header = ['项目'] + selected
                n_cols = len(pl_header)
                # Add change column for 2-period comparison
                if len(selected) == 2:
                    pl_header.append('变动额')
                    pl_header.append('变动率')
                    n_cols = len(pl_header)

                pl_numeric = []
                for fkey, fname in pl_fields:
                    vals = []
                    for s in selected:
                        pl_data = summaries[s].get('pl', {})
                        v = pl_data.get(fkey)
                        vals.append(v)
                    pl_numeric.append((fkey, fname, vals))

                # Render PL table
                pl_table = '| ' + ' | '.join(pl_header) + ' |\n'
                pl_table += '|' + '|'.join(['---'] * n_cols) + '|\n'
                for fkey, fname, vals in pl_numeric:
                    formatted = []
                    for v in vals:
                        if v is None:
                            formatted.append('—')
                        else:
                            formatted.append(f'{v:,.2f}')
                    if len(selected) == 2 and vals[0] is not None and vals[1] is not None:
                        change = vals[0] - vals[1]
                        pct = (change / vals[1] * 100) if vals[1] != 0 else 0
                        change_str = f'{change:+,.2f}'
                        pct_str = f'{pct:+.1f}%'
                        if abs(pct) > 30:
                            change_str += ' ⚠'
                        if abs(pct) > 50:
                            change_str += ' 🔴'
                        formatted.append(change_str)
                        formatted.append(pct_str)
                    pl_table += '| ' + ' | '.join([fname] + formatted) + ' |\n'

                st.markdown(pl_table)

                # ---- BS Comparison ----
                st.subheader('📊 资产负债表对比')

                bs_fields = [
                    ('cash', '货币资金'),
                    ('ar', '应收账款'),
                    ('inventory', '存货'),
                    ('curr_assets', '流动资产合计'),
                    ('fa_net', '固定资产净值'),
                    ('total_assets', '资产总计'),
                    ('ap', '应付账款'),
                    ('tax_payable', '应交税费'),
                    ('other_pay', '其他应付款'),
                    ('total_liab', '负债合计'),
                    ('capital', '实收资本'),
                    ('undist_profit_end', '未分配利润'),
                    ('total_equity', '所有者权益合计'),
                ]

                bs_header = ['项目'] + selected
                if len(selected) == 2:
                    bs_header += ['变动额', '变动率']
                n_bs = len(bs_header)

                bs_table = '| ' + ' | '.join(bs_header) + ' |\n'
                bs_table += '|' + '|'.join(['---'] * n_bs) + '|\n'
                for fkey, fname in bs_fields:
                    vals = []
                    for s in selected:
                        bs_data = summaries[s].get('bs', {})
                        v = bs_data.get(fkey)
                        vals.append(v)
                    formatted = []
                    for v in vals:
                        if v is None:
                            formatted.append('—')
                        else:
                            formatted.append(f'{v:,.2f}')
                    if len(selected) == 2 and vals[0] is not None and vals[1] is not None:
                        change = vals[0] - vals[1]
                        pct = (change / abs(vals[1]) * 100) if vals[1] != 0 else 0
                        fmt = f'{change:+,.2f}'
                        if abs(pct) > 30:
                            fmt += ' ⚠'
                        formatted.append(fmt)
                        formatted.append(f'{pct:+.1f}%')
                    bs_table += '| ' + ' | '.join([fname] + formatted) + ' |\n'

                st.markdown(bs_table)

                # ---- Export to xlsx ----
                st.divider()
                if st.button('📥 导出对比表 (Excel)', use_container_width=True):
                    import openpyxl as _xl
                    from openpyxl.styles import Font as _Font, Alignment as _Align, Border as _B, Side as _S
                    wb = _xl.Workbook()

                    def _write_comparison_sheet(ws, title, fields, summaries_data, selected_periods):
                        ws.title = title
                        thin = _B(left=_S('thin'), right=_S('thin'), top=_S('thin'), bottom=_S('thin'))
                        hdr_font = _Font(name='宋体', size=10, bold=True)
                        num_font = _Font(name='Arial', size=10)
                        # Headers
                        headers = ['项目'] + selected_periods
                        if len(selected_periods) == 2:
                            headers += ['变动额', '变动率']
                        for ci, h in enumerate(headers, 1):
                            c = ws.cell(row=1, column=ci, value=h)
                            c.font = hdr_font; c.border = thin; c.alignment = _Align(horizontal='center')
                        # Data
                        for ri, (fkey, fname) in enumerate(fields, 2):
                            ws.cell(row=ri, column=1, value=fname).font = _Font(name='宋体', size=10)
                            ws.cell(row=ri, column=1).border = thin
                            vals = [summaries_data[s].get(title[:2].lower() if title.startswith('利润') else 'bs', {}).get(fkey) for s in selected_periods]
                            for ci, v in enumerate(vals, 2):
                                c = ws.cell(row=ri, column=ci, value=v)
                                c.font = num_font; c.border = thin; c.number_format = '#,##0.00'
                            if len(selected_periods) == 2 and vals[0] is not None and vals[1] is not None:
                                chg = vals[0] - vals[1]
                                ws.cell(row=ri, column=len(selected_periods)+2, value=chg).font = num_font
                                ws.cell(row=ri, column=len(selected_periods)+2).number_format = '#,##0.00'
                                ws.cell(row=ri, column=len(selected_periods)+2).border = thin
                                pct = chg / abs(vals[1]) * 100 if vals[1] != 0 else 0
                                ws.cell(row=ri, column=len(selected_periods)+3, value=pct/100).font = num_font
                                ws.cell(row=ri, column=len(selected_periods)+3).number_format = '0.0%'
                                ws.cell(row=ri, column=len(selected_periods)+3).border = thin
                        # Col widths
                        ws.column_dimensions[_xl.utils.get_column_letter(1)].width = 18
                        for ci in range(2, len(headers)+2):
                            ws.column_dimensions[_xl.utils.get_column_letter(ci)].width = 16

                    _write_comparison_sheet(wb.active, '利润表对比', pl_fields, summaries, selected)
                    ws2 = wb.create_sheet()
                    _write_comparison_sheet(ws2, '资产负债表对比', bs_fields, summaries, selected)

                    xlsx_path = os.path.join(output_dir, f'{company}_历史对比_{"_".join(selected)}.xlsx')
                    wb.save(xlsx_path)
                    with open(xlsx_path, 'rb') as f:
                        st.download_button('⬇ 下载对比表', f.read(), os.path.basename(xlsx_path),
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True)

                # Auto-highlight anomalies
                if len(selected) == 2:
                    st.divider()
                    st.subheader('🔍 异常波动检测')
                    anomalies = []
                    for fkey, fname in pl_fields + bs_fields:
                        src = 'pl' if fkey in dict(pl_fields) else 'bs'
                        v1 = summaries[selected[0]].get(src, {}).get(fkey)
                        v2 = summaries[selected[1]].get(src, {}).get(fkey)
                        if v1 is not None and v2 is not None and v2 != 0:
                            pct = (v1 - v2) / abs(v2) * 100
                            if abs(pct) > 30:
                                anomalies.append((fname, v1, v2, pct, src))
                    if anomalies:
                        for name, v1, v2, pct, src in anomalies:
                            icon = '🔴' if abs(pct) > 50 else '⚠'
                            direction = '增长' if pct > 0 else '下降'
                            st.warning(
                                f'{icon} **{name}**: {v2:,.2f} → {v1:,.2f}，{direction} {abs(pct):.1f}%'
                            )
                    else:
                        st.success('✅ 未检测到超过30%的异常波动')

                # Trend sparkline hint
                if len(selected) >= 3:
                    st.divider()
                    st.subheader('📈 趋势简析')
                    for fkey, fname in pl_fields[:3] + [pl_fields[-1]]:  # revenue, cogs, surcharges, net_profit
                        vals = [summaries[s].get('pl', {}).get(fkey) for s in selected]
                        vals_clean = [v for v in vals if v is not None]
                        if len(vals_clean) >= 2:
                            trend = '↑ 上升' if vals_clean[-1] > vals_clean[0] else '↓ 下降' if vals_clean[-1] < vals_clean[0] else '→ 持平'
                            st.caption(f'{fname}: {trend}（{vals_clean[0]:,.0f} → {vals_clean[-1]:,.0f}）')

# ---- Footer ----
st.divider()
st.caption(f'数据存储位置：{DATA_DIR}  |  所有数据处理在本地完成，不上传任何服务器')
