# -*- coding: utf-8 -*-
"""科目余额表生成 — 从发票数据汇总各科目借/贷/余额."""
import os
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _sf(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def generate_trial_balance(sales_data, costs_data,
                           company_name='', period_label='',
                           output_path=None):
    """从销售和成本发票数据生成科目余额表.

    Args:
        sales_data: {month: [{account_code, amount, tax_amount, ...}, ...]}
        costs_data: {month: [{account_code, amount, tax_amount, ...}, ...]}
        company_name: 公司全称
        period_label: 期间标签如 '2026年5月'
        output_path: 输出路径，默认 data/output/科目余额表.xlsx

    Returns:
        output_path
    """
    # 汇总每个科目编码
    # 销售发票科目 → 贷方（营业收入类）
    # 成本发票科目 → 借方（成本费用类）
    codes = defaultdict(lambda: {'debit': 0.0, 'debit_tax': 0.0,
                                  'credit': 0.0, 'credit_tax': 0.0,
                                  'goods': set()})

    for mo, records in sales_data.items():
        for rec in records:
            code = rec.get('account_code', '')
            if not code:
                continue
            codes[code]['credit'] += _sf(rec.get('amount', 0))
            codes[code]['credit_tax'] += _sf(rec.get('tax_amount', 0))
            g = rec.get('goods_name', '')
            if g:
                codes[code]['goods'].add(str(g)[:30])

    for mo, records in costs_data.items():
        for rec in records:
            code = rec.get('account_code', '')
            if not code:
                continue
            codes[code]['debit'] += _sf(rec.get('amount', 0))
            codes[code]['debit_tax'] += _sf(rec.get('tax_amount', 0))
            g = rec.get('goods_name', '')
            if g:
                codes[code]['goods'].add(str(g)[:30])

    # Sort by code
    sorted_codes = sorted(codes.items())

    # ---- Create workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '科目余额表'

    # Styles (宋体 like the templates)
    font_title = Font(name='宋体', size=14)
    font_subtitle = Font(name='宋体', size=10)
    font_header = Font(name='宋体', size=10, bold=True)
    font_body = Font(name='宋体', size=10)
    font_number = Font(name='Arial', size=10)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    num_fmt = '#,##0.00'

    NCOLS = 8
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(row=1, column=1, value='科目余额表')
    c.font = font_title
    c.alignment = align_center
    ws.row_dimensions[1].height = 24

    # Row 2: Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c = ws.cell(row=2, column=1, value=f'编制单位：{company_name}    期间：{period_label}    单位：元')
    c.font = font_subtitle
    c.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[2].height = 18

    # Row 3: Column headers
    headers = ['科目编码', '科目名称（参考）', '本期借方', '借方税额',
               '本期贷方', '贷方税额', '余额方向', '期末余额']
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.font = font_header
        c.alignment = align_center
        c.border = thin_border
        c.fill = header_fill
    ws.row_dimensions[3].height = 20

    # Data rows
    for i, (code, vals) in enumerate(sorted_codes):
        row = 4 + i
        debit = round(vals['debit'], 2)
        credit = round(vals['credit'], 2)
        balance = round(credit - debit, 2)
        direction = '贷' if balance >= 0 else '借'

        # 科目名称 — use first goods name as reference
        goods_ref = '、'.join(list(vals['goods'])[:3]) if vals['goods'] else ''

        row_data = [
            code,
            goods_ref,
            debit if debit else None,
            round(vals['debit_tax'], 2) if vals['debit_tax'] else None,
            credit if credit else None,
            round(vals['credit_tax'], 2) if vals['credit_tax'] else None,
            direction,
            abs(balance) if balance != 0 else None,
        ]

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = thin_border
            if ci in (1, 7):  # text columns
                c.font = font_body
                c.alignment = align_left if ci == 1 else align_center
            elif ci == 2:  # goods reference
                c.font = font_body
                c.alignment = align_left
            else:  # number columns
                c.font = font_number
                c.alignment = align_right
                c.number_format = num_fmt

        ws.row_dimensions[row].height = 18

    # Totals row
    total_row = 4 + len(sorted_codes)
    total_debit = round(sum(v['debit'] for _, v in sorted_codes), 2)
    total_credit = round(sum(v['credit'] for _, v in sorted_codes), 2)
    total_debit_tax = round(sum(v['debit_tax'] for _, v in sorted_codes), 2)
    total_credit_tax = round(sum(v['credit_tax'] for _, v in sorted_codes), 2)
    total_balance = round(total_credit - total_debit, 2)

    total_data = [
        '合计', '',
        total_debit if total_debit else None,
        total_debit_tax if total_debit_tax else None,
        total_credit if total_credit else None,
        total_credit_tax if total_credit_tax else None,
        '贷' if total_balance >= 0 else '借',
        abs(total_balance) if total_balance != 0 else None,
    ]

    for ci, val in enumerate(total_data, 1):
        c = ws.cell(row=total_row, column=ci, value=val)
        c.font = Font(name='宋体', size=10, bold=True)
        c.border = thin_border
        c.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        if ci in (1, 7):
            c.alignment = align_center
        elif ci == 2:
            c.alignment = align_left
        else:
            c.alignment = align_right
            c.number_format = num_fmt
    ws.row_dimensions[total_row].height = 20

    # Column widths
    col_widths = [28, 36, 16, 14, 16, 14, 10, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Save
    if output_path is None:
        output_path = str(Path(__file__).parent / 'data' / 'output' / '科目余额表.xlsx')
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
