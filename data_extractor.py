# -*- coding: utf-8 -*-
"""Data extractors — read source Excel files into standardized dict structures.

Three extractors:
  InvoiceExtractor  — sales/cost invoice .xlsx files
  BankExtractor     — bank statement .xls files (农行/信用社)
  PayrollExtractor  — monthly payroll .xlsx files

All extractors are file-path-in, dict-out. No global state. No accounting logic.
"""
import openpyxl
import xlrd
from datetime import datetime
from collections import defaultdict


def _sf(v):
    """Safe float conversion. Returns 0.0 on failure."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _is_date(v):
    """Check if a value looks like a date."""
    if v is None:
        return False
    if isinstance(v, datetime):
        return True
    s = str(v).strip()
    return len(s) >= 10 and s[4] == '-' and s[7] == '-'


def _is_date_xl(v):
    """Check if an xlrd cell value is a date (float serial or string)."""
    if v is None or v == '':
        return False
    if isinstance(v, float) and v > 40000:
        return True
    s = str(v).strip()
    return len(s) >= 10 and s[4] == '-' and s[7] == '-'


# ================================================================
# InvoiceExtractor
# ================================================================

class InvoiceExtractor:
    """Extract invoice data from sales/cost .xlsx files.

    Expected format (金蝶导出):
      - Sheet name: '信息汇总表'
      - Row 1: header
      - Data from row 2
      - Column 8 (0-indexed): date
      - Column 9: 科目编码
      - Column 16: 金额 (不含税)
      - Column 18: 税额

    Returns:
      {month: [{date, invoice_no, account_code, amount, tax_amount}]}
    """

    def __init__(self, filepath, sheet_name='信息汇总表',
                 col_date=8, col_code=9, col_amount=16, col_tax=18, col_goods=11):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.col_date = col_date
        self.col_code = col_code
        self.col_amount = col_amount
        self.col_tax = col_tax
        self.col_goods = col_goods

    def extract(self):
        """Read the file and return ({month_int: [records]}, skipped_count)."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"Sheet '{self.sheet_name}' not found in {self.filepath}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[self.sheet_name]

        result = defaultdict(list)
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = row[self.col_date] if self.col_date < len(row) else None
            if not _is_date(date_val):
                continue

            mo = date_val.month if isinstance(date_val, datetime) else \
                 datetime.strptime(str(date_val)[:10], '%Y-%m-%d').month

            code = str(row[self.col_code]) if self.col_code < len(row) and row[self.col_code] else ''
            amount = _sf(row[self.col_amount]) if self.col_amount < len(row) else 0.0
            tax = _sf(row[self.col_tax]) if self.col_tax < len(row) else 0.0

            if abs(amount) < 0.005 and abs(tax) < 0.005:
                skipped += 1
                continue

            result[mo].append({
                'date': date_val,
                'account_code': code,
                'amount': amount,
                'tax_amount': tax,
                'goods_name': str(row[self.col_goods]) if self.col_goods is not None and self.col_goods < len(row) and row[self.col_goods] else '',
            })

        wb.close()
        return dict(result), skipped

    def get_total_by_code(self, months=None):
        """Extract and return {account_code: total_amount} for convenience.
        Used by the mapping wizard to discover account codes.
        """
        data, _ = self.extract()
        totals = defaultdict(float)
        for mo, records in data.items():
            if months and mo not in months:
                continue
            for rec in records:
                totals[rec['account_code']] += rec['amount']
        return dict(totals)

    def get_sum_by_months(self):
        """Return {month: total_amount} summary."""
        data, _ = self.extract()
        return {mo: sum(r['amount'] for r in records) for mo, records in data.items()}


# ================================================================
# BankExtractor
# ================================================================

class BankExtractor:
    """Extract bank transaction data from .xls files.

    Supports two bank formats:
      - 农行 (Agricultural Bank): chronological order, data from row 3
        col 0=date, 1=收入, 2=支出, 3=余额, 7=摘要
      - 信用社 (Rural Credit Union): reverse chronological, data from row 4
        col 0=date, 1=支出, 2=收入, 3=余额

    Returns:
      {'in_total': float, 'out_total': float, 'end_balance': float,
       'transactions': [{date, amount_in, amount_out, balance, summary}]}
    """

    BANK_NONGHANG = 'nonghang'
    BANK_XINYONGSHE = 'xinyongshe'

    def __init__(self, filepath, bank_type=BANK_NONGHANG):
        self.filepath = filepath
        self.bank_type = bank_type

    def extract(self):
        """Read the bank statement and return standardized dict."""
        wb = xlrd.open_workbook(self.filepath)
        ws = wb.sheet_by_index(0)

        if self.bank_type == self.BANK_NONGHANG:
            return self._extract_nonghang(ws)
        elif self.bank_type == self.BANK_XINYONGSHE:
            return self._extract_xinyongshe(ws)
        else:
            raise ValueError(f"Unknown bank_type: {self.bank_type}")

    def _extract_nonghang(self, ws):
        in_total, out_total = 0.0, 0.0
        end_balance = 0.0
        transactions = []

        for r in range(3, ws.nrows):
            if not _is_date_xl(ws.cell_value(r, 0)):
                continue
            amt_in = _sf(ws.cell_value(r, 1))
            amt_out = _sf(ws.cell_value(r, 2))
            bal = _sf(ws.cell_value(r, 3))
            summary = str(ws.cell_value(r, 7)) if ws.ncols > 7 else ''

            in_total += amt_in
            out_total += amt_out
            end_balance = bal

            transactions.append({
                'date': ws.cell_value(r, 0),
                'amount_in': amt_in,
                'amount_out': amt_out,
                'balance': bal,
                'summary': summary,
            })

        return {
            'in_total': in_total,
            'out_total': out_total,
            'end_balance': end_balance,
            'transactions': transactions,
            'bank_type': self.bank_type,
        }

    def _extract_xinyongshe(self, ws):
        in_total, out_total = 0.0, 0.0
        first_balance = None
        transactions = []

        for r in range(4, ws.nrows):
            if not _is_date_xl(ws.cell_value(r, 0)):
                continue
            amt_out = _sf(ws.cell_value(r, 1))  # Note: reversed columns
            amt_in = _sf(ws.cell_value(r, 2))
            bal = _sf(ws.cell_value(r, 3))

            if first_balance is None:
                first_balance = bal

            in_total += amt_in
            out_total += amt_out

            transactions.append({
                'date': ws.cell_value(r, 0),
                'amount_in': amt_in,
                'amount_out': amt_out,
                'balance': bal,
                'summary': '',
            })

        return {
            'in_total': in_total,
            'out_total': out_total,
            'end_balance': first_balance if first_balance is not None else 0.0,
            'transactions': transactions,
            'bank_type': self.bank_type,
        }

    def get_interest_income(self):
        """Scan transactions for interest income entries (利息 in summary)."""
        data = self.extract()
        total = 0.0
        for t in data['transactions']:
            if '利息' in str(t.get('summary', '')):
                total += t['amount_in']
        return total


# ================================================================
# PayrollExtractor
# ================================================================

class PayrollExtractor:
    """Extract monthly payroll data from .xlsx files.

    Expected format:
      - Row containing '合计' in column 0 marks the summary row
      - Column 5 (0-indexed): 应发工资合计

    Returns:
      {'gross_pay': float}
    """

    def __init__(self, filepath):
        self.filepath = filepath

    def extract(self):
        """Read payroll file and return {'gross_pay': amount}."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb.active

        gross_pay = 0.0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            label = str(row[0]).strip() if row[0] else ''
            if '合计' in label:
                gross_pay = _sf(row[5]) if len(row) > 5 else 0.0
                break

        wb.close()
        return {'gross_pay': gross_pay}
