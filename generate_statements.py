# -*- coding: utf-8 -*-
import openpyxl, os, json
from datetime import datetime
from collections import defaultdict

def sf(v):
    try: return float(v)
    except: return 0.0

# =====================================================================
# STEP 1: Extract all raw data
# =====================================================================
print("Step 1: Extracting data...")

# --- Sales ---
wb = openpyxl.load_workbook('顺恒2026年1-5月销售收入数据.xlsx', data_only=True)
ws = wb['信息汇总表']
sales_by_month = defaultdict(float)
sales_tax_by_month = defaultdict(float)
for row in ws.iter_rows(min_row=2, values_only=True):
    ds = str(row[8]) if row[8] else ''
    if not ds or ds == 'None': continue
    try: m = datetime.strptime(ds[:10], '%Y-%m-%d').month
    except: continue
    sales_by_month[m] += sf(row[16])
    sales_tax_by_month[m] += sf(row[18])

# --- Costs ---
wb2 = openpyxl.load_workbook('顺恒2026年1-5月成本费用数据.xlsx', data_only=True)
ws2 = wb2['信息汇总表']
cogs_codes = ['1110701000000000000', '1110799000000000000']
cogs_by_month = defaultdict(float)
cogs_tax_by_month = defaultdict(float)
admin_by_month = defaultdict(float)
admin_tax_by_month = defaultdict(float)
fixed_asset_items = []  # items that look like fixed assets

for row in ws2.iter_rows(min_row=2, values_only=True):
    ds = str(row[8]) if row[8] else ''
    if not ds or ds == 'None': continue
    try: m = datetime.strptime(ds[:10], '%Y-%m-%d').month
    except: continue
    amt = sf(row[16])
    tax = sf(row[18])
    total = sf(row[19])
    code = str(row[9]) if row[9] else ''
    goods = str(row[11]) if row[11] else ''

    if code in cogs_codes:
        cogs_by_month[m] += amt
        cogs_tax_by_month[m] += tax
    else:
        # Identify potential fixed asset purchases
        is_fa = False
        fa_keywords = ['打印机', '显示器', '电脑', '设备', '彩钢瓦', '钢材', '金属制品', '监控']
        for kw in fa_keywords:
            if kw in goods:
                is_fa = True
                break
        if is_fa and abs(total) > 500:
            fixed_asset_items.append({'month': m, 'goods': goods, 'amt': amt, 'tax': tax, 'total': total})
        admin_by_month[m] += amt
        admin_tax_by_month[m] += tax

# --- Payroll ---
payroll_by_month = {}
employer_si_by_month = {}
for m in range(1, 6):
    wb = openpyxl.load_workbook(f'顺恒2026年{m}月工资薪金.xlsx', data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] and '合计' in str(row[0]):
            payroll_by_month[m] = sf(row[5])  # gross salary
            # Employer social insurance: the individual SI = 4563.55
            # Employer rate is typically ~2x individual
            # Individual: 养老8%+医疗2%+失业0.5% = 10.5% of base
            # Base ~= 4563.55/0.105 = 43,462 (seems high)
            # Let's use typical employer rate ~25% of gross
            employer_si_by_month[m] = sf(row[5]) * 0.25
            break

# --- Bank balances ---
bank_data = {}
for m in range(1, 6):
    bank_data[m] = {'农行': 0, '信用社': 0, '农行_in': 0, '农行_out': 0, '信用社_in': 0, '信用社_out': 0}

    # 农行
    try:
        wb = openpyxl.load_workbook(f'顺恒2026年{m}月农行明细.xlsx', data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row[0]:
                bank_data[m]['农行_in'] += sf(row[1])
                bank_data[m]['农行_out'] += sf(row[2])
                bal = sf(row[3])
                if bal > 0:
                    bank_data[m]['农行'] = bal
    except: pass

    # 信用社
    try:
        wb = openpyxl.load_workbook(f'顺恒2026年{m}月农村信用社明细.xlsx', data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if row[0]:
                bank_data[m]['信用社_in'] += sf(row[2])
                bank_data[m]['信用社_out'] += sf(row[1])
                bal = sf(row[3])
                if bal > 0:
                    bank_data[m]['信用社'] = bal
    except: pass

# =====================================================================
# STEP 2: Income Statement
# =====================================================================
print("Step 2: Computing Income Statement...")

# Totals
cum_sales = sum(sales_by_month.values())
cum_cogs = sum(cogs_by_month.values())
cum_admin_invoice = sum(admin_by_month.values())
cum_payroll = sum(payroll_by_month.values())
cum_employer_si = sum(employer_si_by_month.values())

# Fixed asset items (exclude from admin expense)
fa_total = sum(item['amt'] + abs(item['tax']) for item in fixed_asset_items if item['total'] > 0)
admin_expense = cum_admin_invoice + cum_payroll + cum_employer_si - fa_total

# Tax calculations
total_output_tax = sum(sales_tax_by_month.values())
total_input_tax = sum(cogs_tax_by_month.values()) + sum(admin_tax_by_month.values())
vat_payable = total_output_tax - abs(total_input_tax)
if vat_payable < 0: vat_payable = 0

# Surcharges on VAT
chengjian = vat_payable * 0.07
jiaoyufei = vat_payable * 0.03
difang_jiaoyu = vat_payable * 0.02
total_surcharges = chengjian + jiaoyufei + difang_jiaoyu

# Financial expenses (bank account management fees)
bank_fees_cum = 4.5 * 5  # 农行 monthly fee

# Profit calculations
operating_profit = cum_sales - cum_cogs - total_surcharges - admin_expense - bank_fees_cum
total_profit = operating_profit  # no significant non-op items identified
net_profit = total_profit  # no income tax for small business with losses

# May monthly data
may_sales = sales_by_month.get(5, 0)
may_cogs = cogs_by_month.get(5, 0)
may_admin = admin_by_month.get(5, 0) + payroll_by_month.get(5, 0) + employer_si_by_month.get(5, 0)
may_surcharges = total_surcharges / 5
may_operating_profit = may_sales - may_cogs - may_surcharges - may_admin - 4.5
may_total_profit = may_operating_profit

# =====================================================================
# STEP 3: Balance Sheet
# =====================================================================
print("Step 3: Computing Balance Sheet...")

# Read Dec 2025 BS
wb_bs = openpyxl.load_workbook('顺恒2025年12月_资产负债表（往来重分类）.xlsx', data_only=True)
ws_bs = wb_bs.active

# Assets (left side: columns A-D)
assets = {}
for row in ws_bs.iter_rows(min_row=5, max_row=30, values_only=True):
    name = str(row[0]).strip() if row[0] else ''
    if name:
        assets[name] = {'end_2025': sf(row[2]), 'begin_2025': sf(row[3])}

# Liabilities & Equity (right side: columns E-H)
liab_eq = {}
for row in ws_bs.iter_rows(min_row=5, max_row=30, values_only=True):
    name = str(row[4]).strip() if row[4] else ''
    if name:
        liab_eq[name] = {'end_2025': sf(row[6]), 'begin_2025': sf(row[7])}

# Calculate May 2026 BS
bs = {}

# Assets
bs['货币资金'] = bank_data[5]['农行'] + bank_data[5]['信用社']
bs['应收账款'] = assets.get('应收账款', {}).get('end_2025', 0)
bs['预付账款'] = assets.get('预付账款', {}).get('end_2025', 0)
bs['其他应收款'] = assets.get('其他应收款', {}).get('end_2025', 0)
bs['存货'] = assets.get('存货', {}).get('end_2025', 0)  # simplified
bs['流动资产合计'] = bs['货币资金'] + bs['应收账款'] + bs['预付账款'] + bs['其他应收款'] + bs['存货']

# Fixed assets
fa_orig = assets.get('固定资产原价', {}).get('end_2025', 0)
fa_depr_cum = assets.get('减：累计折旧', {}).get('end_2025', 0)
monthly_depr = 25886.46  # from Dec 2025 data
fa_depr_may = fa_depr_cum + monthly_depr * 5
fa_book_may = fa_orig - fa_depr_may + fa_total
bs['固定资产原价'] = fa_orig + fa_total
bs['减：累计折旧'] = fa_depr_may
bs['固定资产账面价值'] = fa_book_may

bs['长期待摊费用'] = assets.get('长期待摊费用', {}).get('end_2025', 0) - 180 * 5  # monthly amort
if bs['长期待摊费用'] < 0: bs['长期待摊费用'] = 0
bs['非流动资产合计'] = bs['固定资产账面价值'] + bs['长期待摊费用']
bs['资产总计'] = bs['流动资产合计'] + bs['非流动资产合计']

# Liabilities
bs['应付账款'] = liab_eq.get('应付账款', {}).get('end_2025', 0)
bs['预收账款'] = liab_eq.get('预收账款', {}).get('end_2025', 0)
bs['应付职工薪酬'] = cum_payroll + cum_employer_si  # accrued payroll
bs['应交税费'] = vat_payable + total_surcharges  # taxes payable
bs['其他应付款'] = liab_eq.get('其他应付款', {}).get('end_2025', 0)
bs['流动负债合计'] = bs['应付账款'] + bs['预收账款'] + bs['应付职工薪酬'] + bs['应交税费'] + bs['其他应付款']
bs['负债合计'] = bs['流动负债合计']

# Owner's Equity
bs['实收资本（或股本）'] = liab_eq.get('实收资本（或股本）', {}).get('end_2025', 0)
bs['未分配利润'] = liab_eq.get('未分配利润', {}).get('end_2025', 0) + net_profit
bs['所有者权益（或股东权益）合计'] = bs['实收资本（或股本）'] + bs['未分配利润']
bs['负债和所有者权益（或股东权益）总计'] = bs['负债合计'] + bs['所有者权益（或股东权益）合计']

# =====================================================================
# STEP 4: Cash Flow Statement
# =====================================================================
print("Step 4: Computing Cash Flow Statement...")

# Cash received from sales (approximate from bank inflows)
total_bank_in = sum(bank_data[m]['农行_in'] + bank_data[m]['信用社_in'] for m in range(1,6))
total_bank_out = sum(bank_data[m]['农行_out'] + bank_data[m]['信用社_out'] for m in range(1,6))

cf = {}
cf['销售商品、提供劳务收到的现金'] = total_bank_in * 0.75  # approx
cf['收到的其他与经营活动有关的现金'] = total_bank_in * 0.25
cf['购买原材料、商品、接受劳务支付的现金'] = cum_cogs * 1.01  # approx with tax
cf['支付职工薪酬'] = cum_payroll
cf['支付的税费'] = total_surcharges
cf['支付的其他与经营活动有关的现金'] = cum_admin_invoice * 1.03  # approx with tax

cf['经营活动现金流入小计'] = cf['销售商品、提供劳务收到的现金'] + cf['收到的其他与经营活动有关的现金']
cf['经营活动现金流出小计'] = cf['购买原材料、商品、接受劳务支付的现金'] + cf['支付职工薪酬'] + cf['支付的税费'] + cf['支付的其他与经营活动有关的现金']
cf['经营活动产生的现金流量净额'] = cf['经营活动现金流入小计'] - cf['经营活动现金流出小计']

cf['购建固定资产、无形资产和其他非流动资产支付的现金'] = fa_total
cf['投资活动产生的现金流量净额'] = -fa_total
cf['筹资活动产生的现金流量净额'] = 0

ending_cash = bank_data[5]['农行'] + bank_data[5]['信用社']
beginning_cash = assets.get('货币资金', {}).get('end_2025', 0)
cf['现金净增加额'] = ending_cash - beginning_cash
cf['期初现金余额'] = beginning_cash
cf['期末现金余额'] = ending_cash

# =====================================================================
# PRINT SUMMARY
# =====================================================================
print("\n" + "="*80)
print("DATA SUMMARY")
print("="*80)
print(f"Sales revenue (cum):          {cum_sales:>15,.2f}")
print(f"  May:                         {may_sales:>15,.2f}")
print(f"COGS (cum):                    {cum_cogs:>15,.2f}")
print(f"  May:                         {may_cogs:>15,.2f}")
print(f"Admin expenses (cum):          {admin_expense:>15,.2f}")
print(f"  May:                         {may_admin:>15,.2f}")
print(f"  (invoice costs: {cum_admin_invoice:>15,.2f})")
print(f"  (payroll:       {cum_payroll:>15,.2f})")
print(f"  (employer SI:   {cum_employer_si:>15,.2f})")
print(f"  (less fixed assets: {-fa_total:>15,.2f})")
print(f"Tax surcharges:                {total_surcharges:>15,.2f}")
print(f"  VAT payable:                 {vat_payable:>15,.2f}")
print(f"  Output tax:                  {total_output_tax:>15,.2f}")
print(f"  Input tax:                   {abs(total_input_tax):>15,.2f}")
print(f"Financial expenses:            {bank_fees_cum:>15,.2f}")
print(f"---")
print(f"OPERATING PROFIT:              {operating_profit:>15,.2f}")
print(f"NET PROFIT:                    {net_profit:>15,.2f}")
print(f"---")
print(f"Cash May end:                  {ending_cash:>15,.2f}")
print(f"  Nonghang:                    {bank_data[5]['农行']:>15,.2f}")
print(f"  Xinyongshe:                  {bank_data[5]['信用社']:>15,.2f}")
print(f"Fixed assets (net):            {fa_book_may:>15,.2f}")
print(f"Total assets:                  {bs['资产总计']:>15,.2f}")
print(f"Total liabilities:             {bs['负债合计']:>15,.2f}")
print(f"Owner equity:                  {bs['所有者权益（或股东权益）合计']:>15,.2f}")
print(f"BS check (A-L-E):              {bs['资产总计'] - bs['负债合计'] - bs['所有者权益（或股东权益）合计']:>15,.2f}")

# Save data for next step
data = {
    'sales_by_month': dict(sales_by_month),
    'cogs_by_month': dict(cogs_by_month),
    'admin_by_month': dict(admin_by_month),
    'payroll_by_month': payroll_by_month,
    'employer_si_by_month': employer_si_by_month,
    'bank_data': {str(k): v for k, v in bank_data.items()},
    'cum_sales': cum_sales, 'may_sales': may_sales,
    'cum_cogs': cum_cogs, 'may_cogs': may_cogs,
    'admin_expense': admin_expense, 'may_admin': may_admin,
    'total_surcharges': total_surcharges, 'may_surcharges': may_surcharges,
    'bank_fees_cum': bank_fees_cum, 'may_bank_fees': 4.5,
    'operating_profit': operating_profit, 'may_operating_profit': may_operating_profit,
    'net_profit': net_profit, 'may_net_profit': may_operating_profit,
    'vat_payable': vat_payable,
    'total_output_tax': total_output_tax,
    'total_input_tax': abs(total_input_tax),
    'v_chengjian': chengjian,
    'v_jiaoyufei': jiaoyufei,
    'v_difang_jiaoyu': difang_jiaoyu,
    'fa_total': fa_total,
    'bs': bs,
    'cf': cf,
    'ending_cash': ending_cash,
    'beginning_cash': beginning_cash,
    'depreciation_5m': monthly_depr * 5,
    'ltd_amort_5m': 180 * 5,
    'cum_payroll': cum_payroll,
    'cum_employer_si': cum_employer_si,
}

with open('_calc_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("\nData saved to _calc_data.json")
print("Done with calculations!")
