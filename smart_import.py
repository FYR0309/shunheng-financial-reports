# -*- coding: utf-8 -*-
"""Smart file classifier — auto-detects file type by analyzing content structure."""
import os, re
from difflib import SequenceMatcher

import openpyxl

def _fuzzy(a, b):
    return SequenceMatcher(None, a.strip(), b.strip()).ratio()


def _classify_xls(filepath):
    """Fallback classifier for legacy .xls files (using xlrd).

    Handles: bank statements (农行/信用社), balance sheets, profit statements.
    """
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # Read first few rows for header detection
    rows = []
    for r in range(min(10, ws.nrows)):
        rows.append([str(ws.cell_value(r, c)).strip() if ws.cell_value(r, c) else ''
                      for c in range(ws.ncols)])

    # Extract all text from first few rows
    all_text = ' '.join([' '.join(r) for r in rows])

    # Check for bank statement
    if any(r for r in rows if '收入金额' in r or '收入' in r) and \
       any(r for r in rows if '支出金额' in r or '支出' in r):
        # 信用社: check FIRST — 支出 in col 1 and 收入 in col 2 (reversed order)
        for r in rows:
            if len(r) >= 3 and '支出' in r[1] and '收入' in r[2]:
                return 'bank_xin'
        # 农行: has 摘要 or 对方 info
        if '摘要' in all_text or '对方' in all_text:
            return 'bank_nong'
        return 'bank_nong'

    # Check for balance sheet
    if any(kw in all_text for kw in ['资产', '负债']):
        if '期末余额' in all_text or '年初余额' in all_text:
            return 'balance_sheet'

    # Check for profit statement
    if '营业收入' in all_text and '营业成本' in all_text:
        return 'balance_sheet'  # treated as reference, not for import

    return 'unknown'

def classify_file(filepath):
    """Auto-detect what kind of financial file this is.

    Returns one of: 'sales_invoice', 'cost_invoice', 'bank_nong', 'bank_xin',
                    'payroll', 'balance_sheet', 'unknown'
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return 'unknown'

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        # Try xlrd for legacy .xls files
        if ext == '.xls':
            try:
                import xlrd
                return _classify_xls(filepath)
            except Exception:
                return 'unknown'
        return 'unknown'

    sheets = [s.lower() for s in wb.sheetnames]

    # Check for 金蝶 invoice format
    if any('汇总' in s for s in sheets) or any('信息' in s for s in sheets):
        ws = wb.active
        # Read first row headers
        headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header_str = ' '.join(headers)

        # Check for invoice-specific columns
        has_date = any(h for h in headers if '日期' in h)
        has_code = any(h for h in headers if '科目编码' in h or '分类编码' in h or '科目' in h)
        has_amount = any(h for h in headers if '金额' in h)
        has_tax = any(h for h in headers if '税额' in h)
        has_goods = any(h for h in headers if '货品' in h or '货物' in h)

        if has_date and has_code and has_amount:
            # --- Reliable detection: check if col 5 (销方名称) is the same across rows ---
            # Sales file: col 5 = our company (same in every row)
            # Cost file:  col 5 = various suppliers (different per row)
            seller_names = set()
            buyer_names = set()
            for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row), values_only=True):
                if len(row) > 5 and row[5]:
                    seller_names.add(str(row[5]).strip())
                if len(row) > 7 and row[7]:
                    buyer_names.add(str(row[7]).strip())
            # If seller is always the same → sales invoice (one company selling to many)
            # If buyer is always the same → cost invoice (many suppliers selling to one company)
            if len(seller_names) == 1 and len(buyer_names) > 1:
                return 'sales_invoice'
            if len(buyer_names) == 1 and len(seller_names) > 1:
                return 'cost_invoice'

            # --- Fallback: keyword-based detection in goods names ---
            sales_kw = 0
            cost_kw = 0
            for row in ws.iter_rows(min_row=2, max_row=min(30, ws.max_row), values_only=True):
                row_text = ' '.join([str(c) for c in row if c and isinstance(c, str)])
                if any(kw in row_text for kw in ['废铁', '废铜', '废铝', '废钢', '废不锈钢',
                                                   '废电机', '废电线', '报废车辆', '残值']):
                    sales_kw += 1
                if any(kw in row_text for kw in ['采购', '配件', '运费', '维修', '办公', '耗材',
                                                   '修理', '油料', '柴油', '汽油', '物流',
                                                   '加工', '劳务', '服务费', '咨询', '租赁',
                                                   '检测', '拖车', '切割', '拆解']):
                    cost_kw += 1

            if sales_kw > cost_kw:
                return 'sales_invoice'
            elif cost_kw > sales_kw:
                return 'cost_invoice'
            # Fallback: costs typically have many more line items than sales
            row_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if _[0] is not None)
            if row_count > 50:
                return 'cost_invoice'
            return 'sales_invoice'

    # Check for bank statement format — search first 5 rows for headers
    # (Bank files often have headers on row 3-4, not row 1)
    all_rows = []
    for row in wb.active.iter_rows(min_row=1, max_row=min(5, wb.active.max_row), values_only=True):
        all_rows.append([str(c).strip() if c else '' for c in row])
    all_text = ' '.join([' '.join(r) for r in all_rows])

    # Also read up to row 50 for statement detection (利润表/现金流量表 have key terms deeper)
    deep_text = all_text
    for row in wb.active.iter_rows(min_row=6, max_row=min(50, wb.active.max_row), values_only=True):
        deep_text += ' ' + ' '.join([str(c).strip() if c else '' for c in row])

    if '收入' in all_text and '支出' in all_text:
        # Check for 信用社 first: 支出 in col 1, 收入 in col 2
        for r in all_rows:
            if len(r) >= 3 and '支出' in r[1] and '收入' in r[2]:
                return 'bank_xin'
        if '摘要' in all_text or '对方' in all_text:
            return 'bank_nong'
        return 'bank_nong'

    # Check for payroll format — search first 5 rows
    if any('工资' in h or '薪酬' in h or '应发' in h for h in all_text.split()):
        # Check for 合计 row
        for row in wb.active.iter_rows(min_row=1, max_row=min(50, wb.active.max_row), values_only=True):
            first_cell = str(row[0]).strip() if row[0] else ''
            if '合计' in first_cell or '合計' in first_cell:
                return 'payroll'

    # Check for balance sheet format
    asset_kw = sum(1 for kw in ['资产', '负债', '所有者权益', '期末余额', '年初余额'] if kw in all_text)
    if asset_kw >= 2:
        return 'balance_sheet'

    # Check for trial balance (余额表)
    if '科目代码' in all_text and '科目名称' in all_text and '期末余额' in all_text:
        return 'balance_sheet'

    # Check for profit statement (利润表) — reference file (key terms may be deep)
    if '营业收入' in deep_text and '营业成本' in deep_text and '利润总额' in deep_text:
        return 'balance_sheet'

    # Check for cash flow statement (现金流量表) — reference file (key terms may be deep)
    if '经营活动' in deep_text and '投资活动' in deep_text and '筹资活动' in deep_text:
        return 'balance_sheet'

    wb.close()
    return 'unknown'


def parse_opening_balance(filepath):
    """Parse a balance sheet xlsx and extract line items with ending balances.

    Handles the standard 小企业会计准则 balance sheet format:
    - Left side: assets (cols A-D)
    - Right side: liabilities + equity (cols E-H)
    - Returns: {item_name: ending_balance, ...}
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return {}, ['无法打开文件']

    ws = wb.active
    result = {}
    errors = []

    # Read all rows to find the data section
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=True):
        all_rows.append([v for v in row])

    # Detect column structure
    # Standard format: col A=资产, col B=行次, col C=期末余额, col D=年初余额
    #                  col E=负债和所有者权益, col F=行次, col G=期末余额, col H=年初余额
    # But we need to be flexible

    # Find the data start row (skip headers)
    data_start = None
    for i, row in enumerate(all_rows):
        row_str = ' '.join([str(v) for v in row if v])
        if any(kw in row_str for kw in ['货币资金', '应收账款', '短期借款', '实收资本']):
            data_start = i
            break

    if data_start is None:
        # Try to find by scanning for common BS items
        for i, row in enumerate(all_rows):
            first = str(row[0]).strip() if row[0] else ''
            if first and len(first) > 1 and not any(kw in first for kw in ['资产', '负债', '编制', '项目', '单位']):
                data_start = i
                break

    if data_start is None:
        data_start = 3  # Default: data starts from row 4

    # Helper: safe float with comma stripping
    def _to_float(v):
        if v is None: return None
        if isinstance(v, (int, float)): return float(v)
        try: return float(str(v).replace(',', '').replace('，', '').strip())
        except: return None

    # Scan for asset items (left side, cols 0-3)
    # and liability items (right side, cols 4-7 if available)
    has_right_side = len(all_rows[0]) >= 7 if all_rows else False

    for row in all_rows[data_start:]:
        # Left side: asset items
        left_name = str(row[0]).strip() if len(row) > 0 and row[0] else ''
        left_end = None
        # Find the ending balance column (usually col 2 or 3)
        for ci in [2, 3]:
            if len(row) > ci:
                left_end = _to_float(row[ci])
                if left_end is not None:
                    break

        if left_name and left_end is not None and left_name not in ['', 'None']:
            # Skip section headers but keep data items
            # Only skip pure section headers like "资产", "流动资产：", etc.
            is_section = (left_name in ['资产', '负债', '负债和所有者权益', '所有者权益']
                          or left_name.endswith('：') or left_name.endswith(':'))
            if not is_section:
                result[left_name] = left_end

        # Right side: liability/equity items
        if has_right_side and len(row) > 4:
            right_name = str(row[4]).strip() if row[4] else ''
            right_end = None
            for ci in [6, 7]:
                if len(row) > ci:
                    right_end = _to_float(row[ci])
                    if right_end is not None:
                        break

            if right_name and right_end is not None and right_name not in ['', 'None']:
                is_section_r = (right_name in ['资产', '负债', '负债和所有者权益', '所有者权益']
                                or right_name.endswith('：') or right_name.endswith(':'))
                if not is_section_r:
                    result[right_name] = right_end

    wb.close()
    return result, errors


def match_bs_to_config(bs_items, config_fields):
    """Fuzzy-match parsed balance sheet items to config field names.

    Args:
        bs_items: {item_name: value} from parse_opening_balance()
        config_fields: list of expected field names

    Returns:
        {config_field: value, ...} for matched items
        list of unmatched bs_items
    """
    matched = {}
    unmatched = []

    for bs_name, bs_value in bs_items.items():
        best_score = 0
        best_field = None
        for field in config_fields:
            score = _fuzzy(bs_name, field)
            # Exact substring match gets a boost
            if bs_name in field or field in bs_name:
                score += 0.3
            if score > best_score:
                best_score = score
                best_field = field

        if best_score >= 0.55 and best_field:
            matched[best_field] = bs_value
        else:
            unmatched.append((bs_name, bs_value))

    return matched, unmatched
