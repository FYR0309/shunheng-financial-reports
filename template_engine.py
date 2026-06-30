# -*- coding: utf-8 -*-
"""Template engine — loads preset template configs, manages template selection,
and parses custom xlsx templates for data-field mapping.
"""
import json
import os
import re
from difflib import SequenceMatcher

import openpyxl

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'templates')
CUSTOM_DIR = os.path.join(TEMPLATE_DIR, 'custom')
os.makedirs(CUSTOM_DIR, exist_ok=True)

PRESET_MAP = {
    'pl': 'pl_config.json',
    'bs': 'bs_config.json',
    'cf': 'cf_config.json',
}

def _load_json(filename):
    path = os.path.join(TEMPLATE_DIR, filename)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Template file not found: {path}")
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_preset(statement_type):
    """Load a preset template config by type.

    Args:
        statement_type: 'pl' | 'bs' | 'cf'

    Returns:
        dict with keys: template_type, subtitle, ncols, col_widths,
        header_rows, data_start_row, total_rows, columns, value_col_labels,
        merge_ranges, items
    """
    filename = PRESET_MAP.get(statement_type)
    if filename is None:
        raise ValueError(f"Unknown statement type '{statement_type}'. Use 'pl', 'bs', or 'cf'.")
    return _load_json(filename)

def list_presets():
    """Return list of available preset template types."""
    return list(PRESET_MAP.keys())

def get_item_names(statement_type):
    """Return ordered list of item names from a preset template."""
    config = load_preset(statement_type)
    if statement_type == 'bs':
        left = [item['left'] for item in config['items'] if item['left']]
        right = [item['right'] for item in config['items'] if item['right']]
        return left, right
    names = [item['name'] for item in config['items'] if item['name']]
    return names, []


# ================================================================
# Custom template parsing
# ================================================================

def _fuzzy_score(a, b):
    """Return similarity score between two strings (0-1)."""
    a_clean = re.sub(r'[（(].*?[）)]', '', a).strip()
    b_clean = re.sub(r'[（(].*?[）)]', '', b).strip()
    return SequenceMatcher(None, a_clean, b_clean).ratio()


def _detect_statement_type(row_labels):
    """Detect whether a template is PL, BS, or CF based on keyword density."""
    text = ' '.join(row_labels)
    scores = {
        'pl': sum(1 for kw in ['营业收入', '营业成本', '利润', '费用'] if kw in text),
        'bs': sum(1 for kw in ['资产', '负债', '所有者权益', '货币资金', '固定资产'] if kw in text),
        'cf': sum(1 for kw in ['现金流量', '经营活动', '投资活动', '筹资活动'] if kw in text),
    }
    if max(scores.values()) == 0:
        return 'pl'  # default
    return max(scores, key=scores.get)


def parse_custom_template(xlsx_path, statement_type=None, value_columns=(2, 3)):
    """Parse a custom xlsx template and auto-map rows to data fields.

    Reads the first sheet of the xlsx, extracts row labels from column A,
    and fuzzy-matches them against the preset template's known field names.

    Args:
        xlsx_path: path to the xlsx template file
        statement_type: 'pl' | 'bs' | 'cf' (auto-detect if None)
        value_columns: tuple of (cumulative_col, monthly_col) 0-indexed column
                       indices where values should be written

    Returns:
        dict with keys:
          - statement_type: detected/specified type
          - field_map: {field_name: {'row': int, 'col_cum': int, 'col_mon': int,
                                     'label': str, 'score': float}}
          - row_labels: list of (row_num, label_text) for all parsed rows
          - unmatched: list of (row_num, label_text) that didn't match any field
          - sheet_name: name of the parsed sheet
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    sheet_name = ws.title

    # Extract row labels from column A (0-indexed)
    row_labels = []
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None:
            label = str(cell_val).strip()
            if label and len(label) > 1:  # skip single-char noise
                row_labels.append((r, label))

    wb.close()

    # Detect statement type if not specified
    if statement_type is None:
        labels_text = [lbl for _, lbl in row_labels]
        statement_type = _detect_statement_type(labels_text)

    # Get known field names from the preset template
    if statement_type == 'bs':
        known_names, right_names = get_item_names('bs')
        known_names = known_names + right_names
    else:
        known_names, _ = get_item_names(statement_type)

    # Also add supplementary CF items
    if statement_type == 'cf':
        cf_cfg = load_preset('cf')
        extra = [it['name'] for it in cf_cfg['items'] if it['name'] and '：' not in it['name']]
        known_names = list(set(known_names + extra))

    known_names = [n for n in known_names if n]  # filter empty

    # Match each row label to the best known field name
    cum_col, mon_col = value_columns
    field_map = {}
    unmatched = []

    for row_num, label in row_labels:
        best_score = 0
        best_name = None
        for name in known_names:
            score = _fuzzy_score(label, name)
            if score > best_score:
                best_score = score
                best_name = name
        if best_score >= 0.6:
            field_map[best_name] = {
                'row': row_num,
                'col_cum': cum_col + 1,  # 1-indexed for openpyxl
                'col_mon': mon_col + 1,
                'label': label,
                'score': round(best_score, 2),
            }
        else:
            unmatched.append((row_num, label))

    return {
        'statement_type': statement_type,
        'field_map': field_map,
        'row_labels': row_labels,
        'unmatched': unmatched,
        'sheet_name': sheet_name,
    }


def fill_custom_template(template_path, field_map, data_map, output_path,
                          company_name=None, period_label=None):
    """Fill a custom xlsx template with computed data.

    Loads the template xlsx, writes data values into mapped cells,
    and saves the result to output_path. Preserves all original formatting.

    Args:
        template_path: path to the original xlsx template
        field_map: dict from parse_custom_template(), {field_name: {row, col_cum, col_mon}}
        data_map: dict mapping field_name -> {0: cumulative_value, 1: monthly_value}
        output_path: where to save the filled xlsx
        company_name: optional, written to row 1 col A if provided
        period_label: optional, written to row 2 col A if provided

    Returns:
        output_path
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill company info if provided
    if company_name:
        c = ws.cell(row=1, column=1)
        if c.value is None or str(c.value).strip() == '':
            c.value = f'编制单位：{company_name}'
    if period_label:
        c = ws.cell(row=2, column=1)
        if c.value is None or str(c.value).strip() == '':
            c.value = f'期间：{period_label}'

    # Step 1: Clear ALL mapped value cells (remove residual template data)
    for field_name, cell_info in field_map.items():
        ws.cell(row=cell_info['row'], column=cell_info['col_cum']).value = None
        ws.cell(row=cell_info['row'], column=cell_info['col_mon']).value = None

    # Step 2: Fill data values
    for field_name, cell_info in field_map.items():
        if field_name not in data_map:
            continue
        values = data_map[field_name]
        row = cell_info['row']
        # Cumulative value column
        cum_val = values.get(0)
        if cum_val is not None:
            ws.cell(row=row, column=cell_info['col_cum']).value = round(float(cum_val), 2)
        # Monthly value column
        mon_val = values.get(1)
        if mon_val is not None:
            ws.cell(row=row, column=cell_info['col_mon']).value = round(float(mon_val), 2)

    wb.save(output_path)
    return output_path


def save_custom_config(company_name, statement_type, field_map, template_filename):
    """Save a custom template mapping config for a company.

    Args:
        company_name: company identifier
        statement_type: 'pl' | 'bs' | 'cf'
        field_map: the parsed field_map dict
        template_filename: original template filename (for reference)

    Returns path to the saved config file.
    """
    config = {
        'company': company_name,
        'statement_type': statement_type,
        'template_file': template_filename,
        'field_map': field_map,
    }
    path = os.path.join(CUSTOM_DIR, f'{company_name}_{statement_type}_custom.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    return path


def load_custom_config(company_name, statement_type):
    """Load a saved custom template config, or None if not found."""
    path = os.path.join(CUSTOM_DIR, f'{company_name}_{statement_type}_custom.json')
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None
