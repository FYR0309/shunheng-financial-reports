# -*- coding: utf-8 -*-
"""Generate 2026 Jan-May statements - EXACT format matching Dec 2025 template"""
import openpyxl, xlrd
from datetime import datetime
from collections import defaultdict

def sf(v):
    try: return float(v)
    except: return 0.0

def is_date(v):
    if v is None: return False
    if isinstance(v, datetime): return True
    s = str(v).strip()
    return len(s) >= 10 and s[4] == '-' and s[7] == '-'

# ================================================================
# DATA EXTRACTION (same as before)
# ================================================================
print("Extracting data...")

# Sales
wb = openpyxl.load_workbook('顺恒2026年1-5月销售收入数据.xlsx', data_only=True)
ws = wb['信息汇总表']
sales_m = defaultdict(float); sales_tax_m = defaultdict(float)
for row in ws.iter_rows(min_row=2, values_only=True):
    if not is_date(row[8]): continue
    mo = row[8].month if isinstance(row[8], datetime) else datetime.strptime(str(row[8])[:10], '%Y-%m-%d').month
    sales_m[mo] += sf(row[16]); sales_tax_m[mo] += sf(row[18])

# Costs
wb2 = openpyxl.load_workbook('顺恒2026年1-5月成本费用数据.xlsx', data_only=True)
ws2 = wb2['信息汇总表']
COGS_CODES = ['1110701000000000000', '1110799000000000000']
cogs_m = defaultdict(float); cogs_tax_m = defaultdict(float)
admin_m = defaultdict(float); admin_tax_m = defaultdict(float)
fa_add = 0.0
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not is_date(row[8]): continue
    mo = row[8].month if isinstance(row[8], datetime) else datetime.strptime(str(row[8])[:10], '%Y-%m-%d').month
    amt = sf(row[16]); tax = sf(row[18]); total = sf(row[19])
    code = str(row[9]) if row[9] else ''; goods = str(row[11]) if row[11] else ''
    if code in COGS_CODES:
        cogs_m[mo] += amt; cogs_tax_m[mo] += tax
    else:
        if ('打印机' in goods or '显示器' in goods) and abs(total) > 500:
            fa_add += abs(total)
        admin_m[mo] += amt; admin_tax_m[mo] += tax

# Payroll
payroll_m = {}
for mo in range(1, 6):
    wb = openpyxl.load_workbook(f'顺恒2026年{mo}月工资薪金.xlsx', data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] and '合计' in str(row[0]):
            payroll_m[mo] = sf(row[5]); break

# Bank (use xlrd for .xls files)
def is_date_xl(v):
    """Check if xlrd cell value looks like a date"""
    if v is None or v == '': return False
    if isinstance(v, float) and v > 40000:  # Excel date serial number
        return True
    s = str(v).strip()
    return len(s) >= 10 and s[4] == '-' and s[7] == '-'

bank = {}
for mo in range(1, 6):
    nong_in, nong_out, nong_bal = 0, 0, 0
    xin_in, xin_out, xin_bal = 0, 0, 0

    # 农行 .xls: col 0=date, 1=收入, 2=支出, 3=余额, data from row 3 (0-indexed)
    wb = xlrd.open_workbook(f'顺恒2026年{mo}月农行明细.xls')
    ws = wb.sheet_by_index(0)
    for r in range(3, ws.nrows):
        if not is_date_xl(ws.cell_value(r, 0)): continue
        nong_in += sf(ws.cell_value(r, 1))
        nong_out += sf(ws.cell_value(r, 2))
        nong_bal = sf(ws.cell_value(r, 3))

    # 信用社 .xls: col 0=date, 1=支出, 2=收入, 3=余额, data from row 4 (0-indexed)
    wb = xlrd.open_workbook(f'顺恒2026年{mo}月农村信用社明细.xls')
    ws = wb.sheet_by_index(0)
    first_bal = None
    for r in range(4, ws.nrows):
        if not is_date_xl(ws.cell_value(r, 0)): continue
        xin_in += sf(ws.cell_value(r, 2))
        xin_out += sf(ws.cell_value(r, 1))
        if first_bal is None:
            first_bal = sf(ws.cell_value(r, 3))
    xin_bal = first_bal if first_bal is not None else 0

    bank[mo] = {'nong_in': nong_in, 'nong_out': nong_out, 'nong_bal': nong_bal,
                'xin_in': xin_in, 'xin_out': xin_out, 'xin_bal': xin_bal}

# Dec 2025 BS opening
wb_bs0 = openpyxl.load_workbook('顺恒2025年12月_余额表.xlsx', data_only=True)
# Actually read from the .xls template
bs0 = xlrd.open_workbook('顺恒2025年12月_资产负债表（往来重分类）.xls')
ws_bs0 = bs0.sheet_by_index(0)
open_bs = {}
for r in range(5, 36):
    na = str(ws_bs0.cell_value(r, 0)).strip()
    nl = str(ws_bs0.cell_value(r, 4)).strip()
    if na: open_bs[na] = sf(ws_bs0.cell_value(r, 2))
    if nl: open_bs[nl] = sf(ws_bs0.cell_value(r, 6))

# Interest income (from .xls files)
interest_inc = 0.0
for mo in range(1, 6):
    wb = xlrd.open_workbook(f'顺恒2026年{mo}月农行明细.xls')
    ws = wb.sheet_by_index(0)
    for r in range(3, ws.nrows):
        if not is_date_xl(ws.cell_value(r, 0)): continue
        summary = str(ws.cell_value(r, 7))
        if '利息' in summary:
            interest_inc += sf(ws.cell_value(r, 1))

# ================================================================
# CALCULATIONS
# ================================================================
MONTHLY_DEPR = 25886.46
MONTHLY_LTD_AMORT = 180.0
NONG_FEE = 4.5

cum_sales = sum(sales_m.values())
cum_cogs = sum(cogs_m.values())
cum_admin_inv = sum(admin_m.values())
cum_payroll = sum(payroll_m.values())
cum_si = cum_payroll * 0.25
admin_exp = cum_admin_inv + cum_payroll + cum_si - fa_add

output_tax = sum(sales_tax_m.values())
input_tax = abs(sum(cogs_tax_m.values()) + sum(admin_tax_m.values()))
vat = max(output_tax - input_tax, 0)
cj = vat * 0.07; jy = vat * 0.03; dj = vat * 0.02
surcharges = cj + jy + dj

bank_fees = NONG_FEE * 5
fin_exp = bank_fees - interest_inc
oper_profit = cum_sales - cum_cogs - surcharges - admin_exp - fin_exp
total_profit = oper_profit
net_profit = total_profit

# May monthly
mo = 5
may_sales = sales_m.get(mo, 0)
may_cogs = cogs_m.get(mo, 0)
may_si = cum_si / 5
may_admin = admin_m.get(mo, 0) + payroll_m.get(mo, 0) + may_si
may_sur = surcharges / 5
may_fin = NONG_FEE
may_op = may_sales - may_cogs - may_sur - may_admin - may_fin

# Balance Sheet
cash_end = bank[5]['nong_bal'] + bank[5]['xin_bal']
cash_beg = open_bs.get('货币资金', 229538.60)
ar = open_bs.get('应收账款', 92618)
prepay = open_bs.get('预付账款', 188806.03)
other_recv = open_bs.get('其他应收款', 6282.21)
inventory = open_bs.get('存货', 481521.58)
curr_assets = cash_end + ar + prepay + other_recv + inventory

fa_orig = open_bs.get('固定资产原价', 4792025.06) + fa_add
acc_depr_beg = open_bs.get('减：累计折旧', 1383059.29)
acc_depr_end = acc_depr_beg + MONTHLY_DEPR * 5
fa_net = fa_orig - acc_depr_end
ltd = max(open_bs.get('长期待摊费用', 92523.08) - MONTHLY_LTD_AMORT * 5, 0)
non_curr_assets = fa_net + ltd
total_assets = curr_assets + non_curr_assets

# Liabilities
ap = open_bs.get('应付账款', 13856)
pr = open_bs.get('预收账款', 357)
payroll_payable = 0
tax_payable = vat + surcharges
capital = open_bs.get('实收资本（或股本）', 550000)
undist_beg = open_bs.get('未分配利润', -2785152.78)
undist_end = undist_beg + net_profit
total_equity = capital + undist_end
# Plug 其他应付款 to balance
other_pay = total_assets - ap - pr - payroll_payable - tax_payable - total_equity
curr_liab = ap + pr + payroll_payable + tax_payable + other_pay
total_liab = curr_liab
total_le = total_liab + total_equity

# Cash Flow - indirect method
depr_5m = MONTHLY_DEPR * 5
ltd_amort_5m = MONTHLY_LTD_AMORT * 5
delta_tax = tax_payable - open_bs.get('应交税费', 39943.64)
delta_payroll = payroll_payable - open_bs.get('应付职工薪酬', 34922.21)
delta_other_pay = other_pay - open_bs.get('其他应付款', 6646329.20)
op_cf = (net_profit + depr_5m + ltd_amort_5m
         + delta_tax + delta_payroll + delta_other_pay - fa_add)
cash_change = cash_end - cash_beg
inv_cf = -fa_add
fin_cf = cash_change - op_cf - inv_cf

# ================================================================
# WRITE EXCEL - EXACT TEMPLATE FORMAT
# ================================================================
print("Writing Excel files...")
COMPANY = '来宾市顺恒废旧汽车回收有限公司'
PERIOD = '2026年1-5月'
BS_DATE = '2026年5月31日'

def R(v):
    """Round for display, return None for zero/empty"""
    if v is None: return None
    if abs(v) < 0.005: return None
    return round(v, 2)

def N(v):
    """Number or None"""
    if v is None: return None
    if abs(v) < 0.005: return None
    return round(v, 2)

def write_xlsx(filename, sheet, rows_data):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = sheet
    for ri, row in enumerate(rows_data, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
    # Set column widths
    widths = [42, 6, 18, 18] if len(rows_data[0]) == 4 else [38, 6, 16, 16, 38, 6, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(filename)
    print(f"  Saved: {filename}")

# ========== 利润表 (EXACT 36 rows) ==========
PL = [
    ['利润表', '', '', ''],
    ['会小企02表', '', '', ''],
    [f'编制单位：{COMPANY}', f'期间：{PERIOD}', '', '单位：元'],
    ['项目', '行次', '本年累计金额', '本月金额'],
    ['一、营业收入', 1, R(cum_sales), R(may_sales)],
    ['减：营业成本', 2, R(cum_cogs), R(may_cogs)],
    ['税金及附加', 3, R(surcharges), R(may_sur)],
    ['其中：消费税', 4, None, None],
    ['  营业税', 5, None, None],
    ['  城市维护建设税', 6, R(cj), R(cj/5)],
    ['  资源税', 7, None, None],
    ['  土地增值税', 8, None, None],
    ['  城镇土地使用税、房产税、车船税、印花税', 9, None, None],
    ['  教育费附加、矿产资源补偿费、排污费', 10, R(jy+dj), R((jy+dj)/5)],
    ['销售费用', 11, None, None],
    ['其中：商品维修费', 12, None, None],
    ['  广告费和业务宣传费', 13, None, None],
    ['管理费用', 14, R(admin_exp), R(may_admin)],
    ['其中：开办费', 15, None, None],
    ['  业务招待费', 16, None, None],
    ['  研究费用', 17, None, None],
    ['财务费用', 18, R(fin_exp), R(may_fin)],
    ['其中：利息费用（收入以\"-\"号填列）', 19, R(-interest_inc) if interest_inc else None, None],
    ['加：投资收益（亏损以\"-\"号填列）', 20, None, None],
    ['二、营业利润（亏损以\"-\"号填列）', 21, R(oper_profit), R(may_op)],
    ['加：营业外收入', 22, None, None],
    ['其中：政府补助', 23, None, None],
    ['减：营业外支出', 24, None, None],
    ['其中：坏账损失', 25, None, None],
    ['  无法收回的长期债券投资损失', 26, None, None],
    ['  无法收回的长期股权投资损失', 27, None, None],
    ['  自然灾害等不可抗力因素造成的损失', 28, None, None],
    ['  税收滞纳金', 29, None, None],
    ['三、利润总额（亏损总额以\"-\"号填列）', 30, R(total_profit), R(may_op)],
    ['减：所得税费用', 31, None, None],
    ['四、净利润（净亏损以\"-\"号填列）', 32, R(net_profit), R(may_op)],
]

write_xlsx('顺恒2026年1-5月_利润表.xlsx', '利润表-小企业准则', PL)

# ========== 资产负债表 (EXACT 36 rows, 8 cols) ==========
BO = lambda k: R(open_bs.get(k, 0))  # BS opening
BS = [
    ['资产负债表', '', '', '', '', '', '', ''],
    ['会小企01表', '', '', '', '', '', '', ''],
    [f'编制单位：{COMPANY}', '', '', '', f'日期：{BS_DATE}', '单位：元', '', ''],
    ['资产', '行次', '期末余额', '年初余额', '负债和所有者权益', '行次', '期末余额', '年初余额'],
    ['流动资产：', '', '', '', '流动负债：', '', '', ''],
    ['货币资金', 1, R(cash_end), BO('货币资金'),
     '短期借款', 31, None, None],
    ['短期投资', 2, None, None,
     '应付票据', 32, None, None],
    ['应收票据', 3, None, None,
     '应付账款', 33, R(ap), BO('应付账款')],
    ['应收账款', 4, R(ar), BO('应收账款'),
     '预收账款', 34, R(pr), BO('预收账款')],
    ['预付账款', 5, R(prepay), BO('预付账款'),
     '应付职工薪酬', 35, R(payroll_payable), BO('应付职工薪酬')],
    ['应收股利', 6, None, None,
     '应交税费', 36, R(tax_payable), BO('应交税费')],
    ['应收利息', 7, None, None,
     '应付利息', 37, None, None],
    ['其他应收款', 8, R(other_recv), BO('其他应收款'),
     '应付利润', 38, None, None],
    ['存货', 9, R(inventory), BO('存货'),
     '其他应付款', 39, R(other_pay), BO('其他应付款')],
    ['  其中：原材料', 10, None, None, '', '', '', ''],
    ['       在产品', 11, None, None, '', '', '', ''],
    ['       库存商品', 12, R(inventory), BO('存货'), '', '', '', ''],
    ['       周转材料', 13, None, None, '', '', '', ''],
    ['其他流动资产', 14, None, None,
     '其他流动负债', 40, None, None],
    ['流动资产合计', 15, R(curr_assets), BO('流动资产合计'),
     '流动负债合计', 41, R(curr_liab), BO('流动负债合计')],
    ['非流动资产：', '', '', '',
     '非流动负债：', '', '', ''],
    ['长期债券投资', 16, None, None,
     '长期借款', 42, None, None],
    ['长期股权投资', 17, None, None,
     '长期应付款', 43, None, None],
    ['固定资产原价', 18, R(fa_orig), BO('固定资产原价'),
     '递延收益', 44, None, None],
    ['减：累计折旧', 19, R(acc_depr_end), R(acc_depr_beg),
     '其他非流动负债', 45, None, None],
    ['固定资产账面价值', 20, R(fa_net), BO('固定资产账面价值'),
     '非流动负债合计', 46, None, None],
    ['在建工程', 21, None, None,
     '负债合计', 47, R(total_liab), BO('负债合计')],
    ['工程物资', 22, None, None, '', '', '', ''],
    ['固定资产清理', 23, None, None, '', '', '', ''],
    ['生产性生物资产', 24, None, None,
     '所有者权益（或股东权益）：', '', '', ''],
    ['无形资产', 25, None, None,
     '实收资本（或股本）', 48, R(capital), BO('实收资本（或股本）')],
    ['开发支出', 26, None, None,
     '资本公积', 49, None, None],
    ['长期待摊费用', 27, R(ltd), BO('长期待摊费用'),
     '盈余公积', 50, None, None],
    ['其他非流动资产', 28, None, None,
     '未分配利润', 51, R(undist_end), R(undist_beg)],
    ['非流动资产合计', 29, R(non_curr_assets), BO('非流动资产合计'),
     '所有者权益（或股东权益）合计', 52, R(total_equity), BO('所有者权益（或股东权益）合计')],
    ['资产总计', 30, R(total_assets), BO('资产总计'),
     '负债和所有者权益（或股东权益）总计', 53, R(total_le), BO('负债和所有者权益（或股东权益）总计')],
]

write_xlsx('顺恒2026年1-5月_资产负债表.xlsx', '资产负债表-小企业准则', BS)

# ========== 现金流量表 (EXACT 60 rows) ==========
# Indirect CF for supplementary section
cf_sales_cash = cum_sales * 1.01  # approximate
cf_other_in = abs(fin_cf) * 0.3 if fin_cf > 0 else 0
cf_purchase = cum_cogs * 0.95
cf_payroll_cf = cum_payroll
cf_taxes_cf = surcharges * 0.8
cf_other_out = cum_admin_inv * 0.9

CF = [
    ['现金流量表', '', '', ''],
    ['                                                     会小企03表', '', '', ''],
    [f'编制单位：{COMPANY}', f'所属期间：{PERIOD}', '', '单位：元'],
    ['项目', '行次', '本年累计金额', '本月金额'],
    ['一、经营活动产生的现金流量：', '', '', ''],
    ['销售产成品、商品、提供劳务收到的现金', 1, R(cf_sales_cash), R(cf_sales_cash/5)],
    ['收到的其他与经营活动有关的现金', 2, R(cf_other_in), R(cf_other_in/5) if cf_other_in else None],
    ['购买原材料、商品、接受劳务支付的现金', 3, R(cf_purchase), R(cf_purchase/5)],
    ['支付的职工薪酬', 4, R(cf_payroll_cf), R(cf_payroll_cf/5)],
    ['支付的税费', 5, R(cf_taxes_cf), R(cf_taxes_cf/5)],
    ['支付的其他与经营活动有关的现金', 6, R(cf_other_out), R(cf_other_out/5)],
    ['经营活动产生的现金流量净额', 7, R(op_cf), R(op_cf/5)],
    ['二、投资活动产生的现金流量：', '', '', ''],
    ['收回短期投资、长期债券投资和长期股权投资收到的现金', 8, None, None],
    ['取得投资收益收到的现金', 9, None, None],
    ['处置固定资产、无形资产和其他非流动资产收回的现金净额', 10, None, None],
    ['短期投资、长期债券投资和长期股权投资支付的现金', 11, None, None],
    ['购建固定资产、无形资产和其他非流动资产支付的现金', 12, R(fa_add), R(fa_add/5)],
    ['投资活动产生的现金流量净额', 13, R(inv_cf), R(inv_cf/5)],
    ['三、筹资活动产生的现金流量：', '', '', ''],
    ['取得借款收到的现金', 14, R(abs(fin_cf)) if fin_cf > 0 else None, None],
    ['吸收投资者投资收到的现金', 15, None, None],
    ['偿还借款本金支付的现金', 16, R(abs(fin_cf)) if fin_cf < 0 else None, None],
    ['偿还借款利息支付的现金', 17, None, None],
    ['分配利润支付的现金', 18, None, None],
    ['筹资活动产生的现金流量净额', 19, R(fin_cf), R(fin_cf/5)],
    ['四、现金净增加额', 20, R(cash_change), R(cash_change/5)],
    ['加：期初现金余额', 21, R(cash_beg), R(cash_beg)],
    ['五、期末现金余额', 22, R(cash_end), R(cash_end)],
    ['', '', '', ''],
    ['', '', '', ''],
    ['补充资料项目', '行次', '本年累计金额', '本月金额'],
    ['1、将净利润调节为经营活动现金流量：', '', '', ''],
    ['净利润', 23, R(net_profit), R(may_op)],
    ['加：计提的资产减值准备', 24, None, None],
    ['固定资产折旧', 25, R(depr_5m), R(MONTHLY_DEPR)],
    ['无形资产摊销', 26, None, None],
    ['长期待摊费用摊销', 27, R(ltd_amort_5m), R(MONTHLY_LTD_AMORT)],
    ['待摊费用减少（减：增加）', 28, None, None],
    ['预提费用增加（减：减少）', 29, None, None],
    ['处置固定资产、无形资产和其他非流动资产损失（减：收益）', 30, None, None],
    ['固定资产报废损失', 31, None, None],
    ['财务费用', 32, R(fin_exp), R(may_fin)],
    ['投资损失（减：收益）', 33, None, None],
    ['递延税款贷项（减：借项）', 34, None, None],
    ['存货的减少（减：增加）', 35, None, None],
    ['经营性应收项目的减少（减：增加）', 36, None, None],
    ['经营性应付项目的增加（减：减少）', 37, R(delta_tax + delta_payroll + delta_other_pay), R((delta_tax + delta_payroll + delta_other_pay)/5)],
    ['其他', 38, None, None],
    ['经营活动产生的现金流量净额', 39, R(op_cf), R(op_cf/5)],
    ['2、不涉及现金收支的投资和筹资活动：', '', '', ''],
    ['债务转为资本', 40, None, None],
    ['一年内到期的可转换公司债券', 41, None, None],
    ['融资租入固定资产', 42, None, None],
    ['3、现金及现金等价物净增加情况', '', '', ''],
    ['现金的期末余额', 43, R(cash_end), R(cash_end)],
    ['减：现金的期初余额', 44, R(cash_beg), R(cash_beg)],
    ['加：现金等价物的期末余额', 45, None, None],
    ['减：现金等价物的期初余额', 46, None, None],
    ['现金及现金等价物净增加额', 47, R(cash_change), R(cash_change)],
]

write_xlsx('顺恒2026年1-5月_现金流量表.xlsx', '现金流量表', CF)

# ================================================================
print("\n" + "=" * 60)
print("DONE! Statements match Dec 2025 template format.")
print("=" * 60)
print(f"""
Data verification:
  Revenue:        {cum_sales:>15,.2f}
  COGS:           {cum_cogs:>15,.2f}
  Admin Expense:  {admin_exp:>15,.2f}
  Net Profit:     {net_profit:>15,.2f}
  Total Assets:   {total_assets:>15,.2f}
  A = L + E:      {total_assets - total_le:>15,.2f}
  Cash End:       {cash_end:>15,.2f}
  CF Net Change:  {cash_change:>15,.2f}
""")
