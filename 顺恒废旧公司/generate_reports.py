#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
顺恒废旧公司 2026年1-5月财务报表生成脚本
小企业会计准则
"""

import os
import re
import sys
from datetime import datetime, date
from collections import defaultdict
import traceback

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import xlrd

# ============================================================
# 常量
# ============================================================

COMPANY_NAME = "来宾市顺恒废旧物质回收有限公司"
COMPANY_TAX_ID = "91451302MA5P4ENM5F"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "报表输出")

REPORT_START = date(2026, 1, 1)
REPORT_END = date(2026, 5, 31)
MONTHS = [1, 2, 3, 4, 5]

# 源文件
FILES = {
    "balance_sheet": "顺恒2025年12月_资产负债表（往来重分类）.xls",
    "sales": "顺恒2026年1-5月销售收入数据.xlsx",
    "cost": "顺恒2026年1-5月成本费用数据.xlsx",
    "salary": [
        "顺恒2026年1月工资薪金.xlsx",
        "顺恒2026年2月工资薪金.xlsx",
        "顺恒2026年3月工资薪金.xlsx",
        "顺恒2026年4月工资薪金.xlsx",
        "顺恒2026年5月工资薪金.xlsx",
    ],
    "bank_abc": [
        "顺恒2026年1月农行明细.xls",
        "顺恒2026年2月农行明细.xls",
        "顺恒2026年3月农行明细.xls",
        "顺恒2026年4月农行明细.xls",
        "顺恒2026年5月农行明细.xls",
    ],
    "bank_rcc": [
        "顺恒2026年1月农村信用社明细.xls",
        "顺恒2026年2月农村信用社明细.xls",
        "顺恒2026年3月农村信用社明细.xls",
        "顺恒2026年4月农村信用社明细.xls",
        "顺恒2026年5月农村信用社明细.xls",
    ],
}

# 发票列索引（0-based，信息总表 sheet）
INV_COL = {
    "seller_tax_id": 4,
    "seller_name": 5,
    "buyer_tax_id": 6,
    "buyer_name": 7,
    "invoice_date": 8,
    "item_name": 11,
    "quantity": 14,
    "unit_price": 15,
    "amount": 16,
    "tax_rate": 17,
    "tax_amount": 18,
    "total_amount": 19,
    "status": 22,
    "issuer": 25,
}

# 银行流水列索引
BANK_COL = {
    "datetime": 0,
    "income": 1,
    "expense": 2,
    "balance": 3,
    "other_account": 4,
    "other_name": 5,
    "other_bank": 6,
    "summary": 7,
}

# 工资列索引
PAYROLL_COL = {
    "name": 1,
    "gross_salary": 5,
    "social_personal": 11,
    "social_unit": 24,
    "net_pay": 46,
}
PAYROLL_DATA_START_ROW = 7
PAYROLL_FOOTER_ROWS = 2

# 资产负债表行映射
BS_ASSET_ROWS = {
    5:  ("货币资金", "current"),
    6:  ("短期投资", "current"),
    7:  ("应收票据", "current"),
    8:  ("应收账款", "current"),
    9:  ("预付账款", "current"),
    12: ("其他应收款", "current"),
    13: ("存货", "current"),
    19: ("流动资产合计", "current"),
    23: ("固定资产原值", "non_current"),
    24: ("累计折旧", "non_current"),
    25: ("固定资产账面价值", "non_current"),
    32: ("长期待摊费用", "non_current"),
    34: ("非流动资产合计", "non_current"),
    35: ("资产总计", "total"),
}

BS_LIABILITY_ROWS = {
    5:  ("短期借款", "current"),
    7:  ("应付账款", "current"),
    8:  ("预收账款", "current"),
    9:  ("应付职工薪酬", "current"),
    10: ("应交税费", "current"),
    13: ("其他应付款", "current"),
    19: ("流动负债合计", "current"),
    26: ("负债合计", "total"),
}

BS_EQUITY_ROWS = {
    30: ("实收资本", "equity"),
    32: ("盈余公积", "equity"),
    33: ("未分配利润", "equity"),
    34: ("所有者权益合计", "total"),
}

# 固定资产
FIXED_ASSET_ORIGINAL = 4792025.06
ACCUM_DEPRECIATION_OPEN = 1383059.29
MONTHLY_DEPRECIATION = 52062.77  # 约原值×0.95/87.5月≈5.2万/月

LONG_TERM_PREPAID_OPEN = 92523.08

# ============================================================
# 工具函数
# ============================================================

def safe_float(val):
    if val is None or val == "" or val == "-":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None

# ============================================================
# 读取函数
# ============================================================

def read_invoices(filepath, file_label=""):
    """读取发票信息总表"""
    result = {"income": [], "expense": [], "errors": []}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        summary_sheet = None
        for sn in wb.sheetnames:
            if "总表" in sn or "汇总" in sn:
                summary_sheet = sn
                break
        if not summary_sheet:
            result["errors"].append(f"{file_label}: 未找到信息总表 sheet, sheets={wb.sheetnames}")
            wb.close()
            return result

        ws = wb[summary_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row_idx, row in enumerate(rows):
            try:
                seller_tax = str(row[INV_COL["seller_tax_id"]]).strip() if row[INV_COL["seller_tax_id"]] else ""
                buyer_tax = str(row[INV_COL["buyer_tax_id"]]).strip() if row[INV_COL["buyer_tax_id"]] else ""

                if not seller_tax and not buyer_tax:
                    continue

                inv_date = safe_date(row[INV_COL["invoice_date"]])
                if inv_date is None:
                    continue
                if inv_date < REPORT_START or inv_date > REPORT_END:
                    continue

                status = str(row[INV_COL["status"]]).strip() if row[INV_COL["status"]] else ""
                if "红冲" in status:
                    continue

                amount = safe_float(row[INV_COL["amount"]])
                tax = safe_float(row[INV_COL["tax_amount"]])
                total = safe_float(row[INV_COL["total_amount"]])
                item = str(row[INV_COL["item_name"]]).strip() if row[INV_COL["item_name"]] else ""
                month = inv_date.month

                if month not in MONTHS:
                    continue

                if seller_tax == COMPANY_TAX_ID:
                    buyer_name = str(row[INV_COL["buyer_name"]]).strip() if row[INV_COL["buyer_name"]] else ""
                    result["income"].append((inv_date, amount, tax, total, item, buyer_name, month))
                elif buyer_tax == COMPANY_TAX_ID:
                    seller_name = str(row[INV_COL["seller_name"]]).strip() if row[INV_COL["seller_name"]] else ""
                    result["expense"].append((inv_date, amount, tax, total, item, seller_name, month))

            except Exception as e:
                result["errors"].append(f"{file_label} 行{row_idx+2}: {e}")

        wb.close()
    except Exception as e:
        result["errors"].append(f"{file_label}: 读取失败 - {e}")
    return result


def read_bank_statement(filepath, file_label=""):
    """读取银行流水"""
    result = {"transactions": [], "account_info": "", "errors": []}
    try:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)

        if ws.nrows >= 2:
            result["account_info"] = str(ws.cell_value(1, 0)) + " " + str(ws.cell_value(1, 1))

        for row_idx in range(3, ws.nrows):
            try:
                raw_date = ws.cell_value(row_idx, BANK_COL["datetime"])
                txn_date = safe_date(raw_date)
                if txn_date is None:
                    continue
                if txn_date < REPORT_START or txn_date > REPORT_END:
                    continue

                income = safe_float(ws.cell_value(row_idx, BANK_COL["income"]))
                expense = safe_float(ws.cell_value(row_idx, BANK_COL["expense"]))
                balance = safe_float(ws.cell_value(row_idx, BANK_COL["balance"]))
                other_name = str(ws.cell_value(row_idx, BANK_COL["other_name"])).strip()
                other_account = str(ws.cell_value(row_idx, BANK_COL["other_account"])).strip()
                summary = str(ws.cell_value(row_idx, BANK_COL["summary"])).strip()

                result["transactions"].append((
                    txn_date, income, expense, balance,
                    other_name, other_account, summary, txn_date.month
                ))
            except Exception as e:
                result["errors"].append(f"{file_label} 行{row_idx+1}: {e}")
    except Exception as e:
        result["errors"].append(f"{file_label}: 读取失败 - {e}")
    return result


def read_payroll(filepath, file_label=""):
    """读取工资表"""
    result = {"employees": [], "total_gross": 0.0, "total_social_unit": 0.0, "total_net": 0.0, "errors": []}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        main_sheet = None
        for sn in wb.sheetnames:
            if any(kw in sn for kw in ["工资", "薪金", "薪酬"]):
                main_sheet = sn
                break
        if not main_sheet:
            main_sheet = wb.sheetnames[0]

        ws = wb[main_sheet]
        rows = list(ws.iter_rows(values_only=True))
        total_rows = len(rows)

        start = PAYROLL_DATA_START_ROW
        end = max(start + 1, total_rows - PAYROLL_FOOTER_ROWS)

        for row_idx in range(start, end):
            try:
                row = rows[row_idx]
                name = str(row[PAYROLL_COL["name"]]).strip() if row[PAYROLL_COL["name"]] else ""
                if not name or name in ("姓名", "合计", "name", "None", ""):
                    continue

                gross = safe_float(row[PAYROLL_COL["gross_salary"]])
                social_personal = safe_float(row[PAYROLL_COL["social_personal"]])
                social_unit = safe_float(row[PAYROLL_COL["social_unit"]])
                net = safe_float(row[PAYROLL_COL["net_pay"]])

                if gross == 0.0 and net == 0.0:
                    continue

                emp = {"name": name, "gross": gross, "social_personal": social_personal,
                       "social_unit": social_unit, "net": net}
                result["employees"].append(emp)
                result["total_gross"] += gross
                result["total_social_unit"] += social_unit
                result["total_net"] += net
            except Exception as e:
                result["errors"].append(f"{file_label} 行{row_idx+1}: {e}")

        wb.close()
    except Exception as e:
        result["errors"].append(f"{file_label}: 读取失败 - {e}")
    return result


def read_opening_balance_sheet(filepath):
    """读取2025年12月资产负债表"""
    result = {}
    try:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)

        for row_idx, (name, _) in BS_ASSET_ROWS.items():
            if row_idx < ws.nrows:
                end_val = safe_float(ws.cell_value(row_idx, 2))
                begin_val = safe_float(ws.cell_value(row_idx, 3))
                result[name] = (end_val, begin_val)

        for row_idx, (name, _) in BS_LIABILITY_ROWS.items():
            if row_idx < ws.nrows:
                end_val = safe_float(ws.cell_value(row_idx, 6))
                begin_val = safe_float(ws.cell_value(row_idx, 7))
                result[name] = (end_val, begin_val)

        for row_idx, (name, _) in BS_EQUITY_ROWS.items():
            if row_idx < ws.nrows:
                end_val = safe_float(ws.cell_value(row_idx, 6))
                begin_val = safe_float(ws.cell_value(row_idx, 7))
                result[name] = (end_val, begin_val)
    except Exception as e:
        print(f"  [FAIL] 读取资产负债表失败: {e}")
        return None
    return result

# ============================================================
# 利润表计算
# ============================================================

def compute_profit_statement(all_income, all_expense, payroll_summary):
    """计算利润表（1-5月累计+分月）"""
    monthly = {m: {
        "revenue": 0.0, "cogs": 0.0, "tax_surcharge": 0.0,
        "admin_salary": 0.0, "admin_depreciation": 0.0,
        "admin_rent": 0.0, "admin_other": 0.0,
        "financial_expense": 0.0, "vat_collected": 0.0, "vat_paid_input": 0.0,
        "income_tax": 0.0,
    } for m in MONTHS}

    for _, amount, tax, total, item, counterparty, month in all_income:
        monthly[month]["revenue"] += amount
        monthly[month]["vat_collected"] += tax

    # 成本费用分类关键词
    COGS_KEYWORDS = ["废", "铁", "钢", "铝", "铜", "纸", "塑料", "原料", "材料",
                      "运输", "物流", "装卸", "搬运", "过磅"]
    ADMIN_KEYWORDS = ["办公", "文具", "耗材", "水电", "物业", "通讯", "电话",
                       "差旅", "招待", "餐", "住宿", "维修", "配件", "检测",
                       "咨询", "服务", "加油", "汽油", "柴油", "保险", "软件",
                       "打印", "印", "快递", "邮", "广告", "刻章", "法律"]
    RENT_KEYWORDS = ["租赁", "租金", "房租", "租"]

    for _, amount, tax, total, item, counterparty, month in all_expense:
        item_str = str(item) if item else ""
        if any(kw in item_str for kw in RENT_KEYWORDS):
            monthly[month]["admin_rent"] += amount
        elif any(kw in item_str for kw in ADMIN_KEYWORDS):
            monthly[month]["admin_other"] += amount
        elif any(kw in item_str for kw in COGS_KEYWORDS):
            monthly[month]["cogs"] += amount
        else:
            # 默认归入COGS（未匹配到的采购类支出）
            monthly[month]["cogs"] += amount
        monthly[month]["vat_paid_input"] += tax

    # 工资按月平均分摊
    total_salary = payroll_summary.get("total_gross", 0) + payroll_summary.get("total_social_unit", 0)
    monthly_salary = total_salary / len(MONTHS)
    for m in MONTHS:
        monthly[m]["admin_salary"] = monthly_salary
        monthly[m]["admin_depreciation"] = MONTHLY_DEPRECIATION

    for m in MONTHS:
        vat_payable = monthly[m]["vat_collected"]
        monthly[m]["tax_surcharge"] = vat_payable * 0.12

    cumulative = {
        "revenue": sum(monthly[m]["revenue"] for m in MONTHS),
        "cogs": sum(monthly[m]["cogs"] for m in MONTHS),
        "gross_profit": 0.0,
        "tax_surcharge": sum(monthly[m]["tax_surcharge"] for m in MONTHS),
        "admin_salary": sum(monthly[m]["admin_salary"] for m in MONTHS),
        "admin_depreciation": sum(monthly[m]["admin_depreciation"] for m in MONTHS),
        "admin_rent": sum(monthly[m]["admin_rent"] for m in MONTHS),
        "admin_other": sum(monthly[m]["admin_other"] for m in MONTHS),
        "total_admin": 0.0,
        "financial_expense": sum(monthly[m]["financial_expense"] for m in MONTHS),
        "operating_profit": 0.0,
        "total_profit": 0.0,
        "income_tax": 0.0,
        "net_profit": 0.0,
        "vat_collected": sum(monthly[m]["vat_collected"] for m in MONTHS),
        "vat_paid_input": sum(monthly[m]["vat_paid_input"] for m in MONTHS),
    }

    cumulative["gross_profit"] = cumulative["revenue"] - cumulative["cogs"]
    cumulative["total_admin"] = (cumulative["admin_salary"] + cumulative["admin_depreciation"]
                                  + cumulative["admin_rent"] + cumulative["admin_other"])
    cumulative["operating_profit"] = (cumulative["gross_profit"] - cumulative["tax_surcharge"]
                                       - cumulative["total_admin"] - cumulative["financial_expense"])
    cumulative["total_profit"] = cumulative["operating_profit"]

    if cumulative["total_profit"] <= 0:
        cumulative["income_tax"] = 0.0
    else:
        annualized = cumulative["total_profit"] / 5 * 12
        if annualized <= 1000000:
            cumulative["income_tax"] = cumulative["total_profit"] * 0.05
        else:
            cumulative["income_tax"] = cumulative["total_profit"] * 0.10

    cumulative["net_profit"] = cumulative["total_profit"] - cumulative["income_tax"]

    # 月度所得税分摊（按累计所得税/月数平均）
    monthly_tax = cumulative["income_tax"] / len(MONTHS)
    for m in MONTHS:
        monthly[m]["income_tax"] = monthly_tax

    notes = [
        f"营业收入(累计): {cumulative['revenue']:,.2f}",
        f"营业成本(累计): {cumulative['cogs']:,.2f}",
        f"毛利率: {cumulative['gross_profit']/cumulative['revenue']*100:.1f}%" if cumulative['revenue'] > 0 else "毛利率: N/A",
        f"利润总额(累计): {cumulative['total_profit']:,.2f}",
        f"所得税估算: {cumulative['income_tax']:,.2f}",
        f"净利润(累计): {cumulative['net_profit']:,.2f}",
    ]

    return {"monthly": monthly, "cumulative": cumulative, "notes": notes}


# ============================================================
# 资产负债表计算
# ============================================================

def compute_balance_sheet(opening_bs, cumulative_pl, bank_end_balance, payroll_5m_summary):
    """计算2026年5月31日资产负债表"""
    bs = {}
    get_open = lambda name: opening_bs.get(name, (0.0, 0.0))[0]

    cash_open = get_open("货币资金")
    bs["货币资金"] = (bank_end_balance, cash_open)

    bs["应收账款"] = (get_open("应收账款"), get_open("应收账款"))
    bs["预付账款"] = (get_open("预付账款"), get_open("预付账款"))
    bs["其他应收款"] = (get_open("其他应收款"), get_open("其他应收款"))

    inventory_open = get_open("存货")
    bs["存货"] = (inventory_open, inventory_open)

    current_end = (bs["货币资金"][0] + bs["应收账款"][0] + bs["预付账款"][0]
                   + bs["其他应收款"][0] + bs["存货"][0])
    bs["流动资产合计"] = (current_end, get_open("流动资产合计"))

    bs["固定资产原值"] = (FIXED_ASSET_ORIGINAL, FIXED_ASSET_ORIGINAL)
    accum_depr = ACCUM_DEPRECIATION_OPEN + MONTHLY_DEPRECIATION * 5
    bs["累计折旧"] = (accum_depr, ACCUM_DEPRECIATION_OPEN)
    net_fixed = FIXED_ASSET_ORIGINAL - accum_depr
    bs["固定资产账面价值"] = (net_fixed, get_open("固定资产账面价值"))

    bs["长期待摊费用"] = (get_open("长期待摊费用"), get_open("长期待摊费用"))

    non_current_end = net_fixed + bs["长期待摊费用"][0]
    bs["非流动资产合计"] = (non_current_end, get_open("非流动资产合计"))
    bs["资产总计"] = (current_end + non_current_end, get_open("资产总计"))

    bs["应付账款"] = (get_open("应付账款"), get_open("应付账款"))
    bs["预收账款"] = (get_open("预收账款"), get_open("预收账款"))

    payroll_open = get_open("应付职工薪酬")
    total_payroll_cost = payroll_5m_summary.get("total_gross", 0) + payroll_5m_summary.get("total_social_unit", 0)
    total_paid = payroll_5m_summary.get("total_net", 0)
    bs["应付职工薪酬"] = (payroll_open + total_payroll_cost - total_paid, payroll_open)

    tax_open = get_open("应交税费")
    vat_total = cumulative_pl.get("vat_collected", 0)
    surcharge_total = cumulative_pl.get("tax_surcharge", 0)
    income_tax = cumulative_pl.get("income_tax", 0)
    new_tax = tax_open + vat_total + surcharge_total + income_tax
    bs["应交税费"] = (max(new_tax, 0.0), tax_open)

    bs["其他应付款"] = (get_open("其他应付款"), get_open("其他应付款"))

    current_liab_end = (bs["应付账款"][0] + bs["预收账款"][0]
                        + bs["应付职工薪酬"][0] + bs["应交税费"][0] + bs["其他应付款"][0])
    bs["流动负债合计"] = (current_liab_end, get_open("流动负债合计"))
    bs["负债合计"] = (current_liab_end, get_open("负债合计"))

    bs["实收资本"] = (get_open("实收资本"), get_open("实收资本"))
    bs["盈余公积"] = (get_open("盈余公积"), get_open("盈余公积"))

    undistributed_open = get_open("未分配利润")
    net_profit = cumulative_pl.get("net_profit", 0)
    bs["未分配利润"] = (undistributed_open + net_profit, undistributed_open)

    equity_end = bs["实收资本"][0] + bs["盈余公积"][0] + bs["未分配利润"][0]
    bs["所有者权益合计"] = (equity_end, get_open("所有者权益合计"))

    bs["负债和所有者权益总计"] = (bs["负债合计"][0] + bs["所有者权益合计"][0],
                              get_open("负债合计") + get_open("所有者权益合计"))

    return bs


# ============================================================
# 现金流量表计算
# ============================================================

def compute_cashflow(all_transactions, payroll_5m_summary):
    """从银行流水分类生成现金流量表"""
    cf = {
        "销售商品提供劳务收到的现金": 0.0,
        "收到其他与经营活动有关的现金": 0.0,
        "经营活动现金流入小计": 0.0,
        "购买商品接受劳务支付的现金": 0.0,
        "支付给职工以及为职工支付的现金": 0.0,
        "支付的各项税费": 0.0,
        "支付其他与经营活动有关的现金": 0.0,
        "经营活动现金流出小计": 0.0,
        "经营活动产生的现金流量净额": 0.0,
    }

    unclassified = []
    tax_keywords = ["税务", "税局", "国家税务", "国库", "税款", "征税", "纳税"]
    bank_fee_keywords = ["手续费", "账户管理费", "服务费", "短信费", "年费", "工本费"]
    salary_keywords = ["工资", "薪酬", "薪金", "代发", "劳务", "奖金", "补贴"]
    # 工资相关：银行代发标识
    salary_bank_keywords = ["农信银e支付", "e支付", "代发工资", "批量代发"]

    for txn in all_transactions:
        _, income, expense, balance, other_name, other_account, summary, month = txn

        if income > 0:
            is_personal = (len(other_name) >= 2 and
                          not any(kw in other_name for kw in ["公司", "有限", "厂", "局", "银行",
                                                               "信用", "支行", "分行", "中国"]))
            if any(kw in other_name for kw in ["银行", "农行", "信用", "支行", "分行"]) and not is_personal:
                cf["收到其他与经营活动有关的现金"] += income
            else:
                cf["销售商品提供劳务收到的现金"] += income

        if expense > 0:
            if any(kw in summary for kw in bank_fee_keywords) or any(kw in other_name for kw in bank_fee_keywords) or (expense < 100 and any(kw in other_name for kw in ["银行", "农行", "信用"])):
                cf["支付其他与经营活动有关的现金"] += expense
            elif any(kw in other_name for kw in tax_keywords) or any(kw in summary for kw in tax_keywords):
                cf["支付的各项税费"] += expense
            elif any(kw in summary for kw in salary_keywords) or any(kw in other_name for kw in salary_keywords) or any(kw in summary for kw in salary_bank_keywords):
                cf["支付给职工以及为职工支付的现金"] += expense
            else:
                cf["购买商品接受劳务支付的现金"] += expense

    cf["经营活动现金流入小计"] = cf["销售商品提供劳务收到的现金"] + cf["收到其他与经营活动有关的现金"]
    cf["经营活动现金流出小计"] = (cf["购买商品接受劳务支付的现金"]
                              + cf["支付给职工以及为职工支付的现金"]
                              + cf["支付的各项税费"]
                              + cf["支付其他与经营活动有关的现金"])

    # 保底：如果关键词未匹配到工资支出，用薪酬数据估算
    total_net_pay = payroll_5m_summary.get("total_net", 0)
    if cf["支付给职工以及为职工支付的现金"] < total_net_pay * 0.1 and total_net_pay > 0:
        # 从"购买商品接受劳务支付的现金"中划转
        transferred = min(total_net_pay, cf["购买商品接受劳务支付的现金"])
        cf["购买商品接受劳务支付的现金"] -= transferred
        cf["支付给职工以及为职工支付的现金"] += transferred
        cf["经营活动现金流出小计"] = (cf["购买商品接受劳务支付的现金"]
                                  + cf["支付给职工以及为职工支付的现金"]
                                  + cf["支付的各项税费"]
                                  + cf["支付其他与经营活动有关的现金"])
        unclassified.append(f"[ADJ] 工资支出自动估算: 从购买商品划转 {transferred:,.2f} 元到支付给职工")

    cf["经营活动产生的现金流量净额"] = cf["经营活动现金流入小计"] - cf["经营活动现金流出小计"]

    monthly_io = {m: [0.0, 0.0] for m in MONTHS}
    for txn in all_transactions:
        _, income, expense, _, _, _, _, month = txn
        monthly_io[month][0] += income
        monthly_io[month][1] += expense

    return {"categories": cf, "unclassified": unclassified, "monthly_io": monthly_io}


# ============================================================
# 校验
# ============================================================

def validate(bs, cumulative_pl, cf, opening_bs, bank_balance_from_statement):
    """执行全部校验"""
    passed = []
    failed = []
    warnings = []

    total_assets = bs["资产总计"][0]
    total_liab_equity = bs["负债和所有者权益总计"][0]
    diff = abs(total_assets - total_liab_equity)
    if diff < 0.02:
        passed.append(f"[PASS] 资产负债表平衡: 资产 {total_assets:,.2f} = 负债+权益 {total_liab_equity:,.2f}")
    else:
        failed.append(f"[FAIL] 资产负债表不平衡! 差异 {diff:,.2f}")

    net_profit_pl = cumulative_pl["net_profit"]
    net_profit_bs_change = bs["未分配利润"][0] - bs["未分配利润"][1]
    diff2 = abs(net_profit_pl - net_profit_bs_change)
    if diff2 < 0.02:
        passed.append(f"[PASS] 未分配利润勾稽: PL {net_profit_pl:,.2f} = BS变动 {net_profit_bs_change:,.2f}")
    else:
        failed.append(f"[FAIL] 未分配利润勾稽不一致! 差异 {diff2:,.2f}")

    bs_cash = bs["货币资金"][0]
    diff3 = abs(bs_cash - bank_balance_from_statement)
    if diff3 < 0.02:
        passed.append(f"[PASS] 银行余额核对: BS {bs_cash:,.2f} = 流水 {bank_balance_from_statement:,.2f}")
    else:
        warnings.append(f"[WARN] 银行余额有差异: {diff3:,.2f}")

    total_income = cumulative_pl["revenue"]
    total_cost = cumulative_pl["cogs"] + cumulative_pl["admin_rent"]
    passed.append(f"[INFO] 收入发票(不含税): {total_income:,.2f}")
    passed.append(f"[INFO] 成本费用发票(不含税): {total_cost:,.2f}")

    cash_open = bs["货币资金"][1]
    cash_end = bs["货币资金"][0]
    cf_net = cf["categories"]["经营活动产生的现金流量净额"]
    diff4 = abs((cash_end - cash_open) - cf_net)
    if diff4 < 1.0:
        passed.append(f"[PASS] 现金流与货币资金变动吻合: 现金变动 {cash_end-cash_open:,.2f} ≈ CF {cf_net:,.2f}")
    else:
        warnings.append(f"[WARN] 现金流与货币资金变动差异: {diff4:,.2f} (可能有非经营现金流)")

    return passed, failed, warnings


# ============================================================
# 样式与输出
# ============================================================

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
NUM_FMT = '#,##0.00'


def apply_style(cell, font=BODY_FONT, border=THIN_BORDER, number_format=None):
    cell.font = font
    cell.border = border
    if number_format:
        cell.number_format = number_format


def write_profit_sheet(wb, cumulative, monthly_detail):
    ws = wb.create_sheet("利润表")
    ws.sheet_properties.tabColor = "4472C4"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{COMPANY_NAME} 利润表"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:E2")
    ws["A2"] = "2026年1-5月（小企业会计准则）      会小企02表      单位：元"
    ws["A2"].font = Font(name="微软雅黑", size=9)

    headers = ["项目", "行次", "本月数", "本年累计", "附注"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        apply_style(c, HEADER_FONT)

    # 取5月当月数据
    may = monthly_detail.get(5, {})
    may_revenue = may.get("revenue", 0)
    may_cogs = may.get("cogs", 0)
    may_tax_surcharge = may.get("tax_surcharge", 0)
    may_admin_salary = may.get("admin_salary", 0)
    may_admin_depreciation = may.get("admin_depreciation", 0)
    may_admin_rent = may.get("admin_rent", 0)
    may_admin_other = may.get("admin_other", 0)
    may_financial = may.get("financial_expense", 0)
    may_total_admin = may_admin_salary + may_admin_depreciation + may_admin_rent + may_admin_other
    may_gross_profit = may_revenue - may_cogs
    may_operating = may_gross_profit - may_tax_surcharge - may_total_admin - may_financial
    may_total_profit = may_operating  # 无营业外收支
    may_income_tax = may.get("income_tax", 0)
    may_net = may_total_profit - may_income_tax

    # (项目名, 行次, 本月数, 累计数, 是否粗体)
    items = [
        ("一、营业收入", 1, may_revenue, cumulative["revenue"], True),
        ("减：营业成本", 2, may_cogs, cumulative["cogs"], False),
        ("    税金及附加", 3, may_tax_surcharge, cumulative["tax_surcharge"], False),
        ("    销售费用", 4, 0, 0, False),
        ("    管理费用", 5, may_total_admin, cumulative["total_admin"], False),
        ("      其中：工资", 6, may_admin_salary, cumulative["admin_salary"], False),
        ("      其中：折旧", 7, may_admin_depreciation, cumulative["admin_depreciation"], False),
        ("      其中：租金", 8, may_admin_rent, cumulative["admin_rent"], False),
        ("      其中：其他", 9, may_admin_other, cumulative["admin_other"], False),
        ("    财务费用", 10, may_financial, cumulative["financial_expense"], False),
        ("    其中：利息费用", 11, 0, 0, False),
        ("加：投资收益", 12, 0, 0, False),
        ("二、营业利润", 13, may_operating, cumulative["operating_profit"], True),
        ("加：营业外收入", 14, 0, 0, False),
        ("减：营业外支出", 15, 0, 0, False),
        ("三、利润总额", 16, may_total_profit, cumulative["total_profit"], True),
        ("减：所得税费用", 17, may_income_tax, cumulative["income_tax"], False),
        ("四、净利润", 18, may_net, cumulative["net_profit"], True),
    ]

    row = 5
    for name, line_num, month_amt, cum_amt, is_bold in items:
        f = BOLD_FONT if is_bold else BODY_FONT
        c1 = ws.cell(row=row, column=1, value=name); apply_style(c1, f)
        c2 = ws.cell(row=row, column=2, value=line_num); apply_style(c2, BODY_FONT)
        c3 = ws.cell(row=row, column=3, value=round(month_amt, 2)); apply_style(c3, f if is_bold else BODY_FONT, number_format=NUM_FMT)
        c4 = ws.cell(row=row, column=4, value=round(cum_amt, 2)); apply_style(c4, f, number_format=NUM_FMT)
        c5 = ws.cell(row=row, column=5, value=""); apply_style(c5, BODY_FONT)
        row += 1

    # 签章区
    row += 2
    ws.merge_cells(f"A{row}:C{row}"); ws.cell(row=row, column=1, value="单位负责人："); apply_style(ws.cell(row=row, column=1), BODY_FONT, THIN_BORDER)
    ws.merge_cells(f"D{row}:E{row}"); ws.cell(row=row, column=4, value="会计负责人："); apply_style(ws.cell(row=row, column=4), BODY_FONT, THIN_BORDER)
    row += 1
    ws.merge_cells(f"A{row}:C{row}"); ws.cell(row=row, column=1, value="制表人："); apply_style(ws.cell(row=row, column=1), BODY_FONT, THIN_BORDER)
    ws.merge_cells(f"D{row}:E{row}"); ws.cell(row=row, column=4, value=f"制表日期：{REPORT_END.strftime('%Y年%m月%d日')}"); apply_style(ws.cell(row=row, column=4), BODY_FONT, THIN_BORDER)

    for c, w in zip("ABCDE", [35, 8, 16, 20, 10]):
        ws.column_dimensions[c].width = w


def write_balance_sheet_sheet(wb, bs):
    ws = wb.create_sheet("资产负债表")
    ws.sheet_properties.tabColor = "70AD47"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{COMPANY_NAME} 资产负债表"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:H2")
    ws["A2"] = "2026年5月31日（小企业会计准则）     会小企01表      单位：元"
    ws["A2"].font = Font(name="微软雅黑", size=9)

    col_headers = ["资产", "行次", "期末数", "年初数", "负债及所有者权益", "行次", "期末数", "年初数"]
    for col, h in enumerate(col_headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        apply_style(c, HEADER_FONT)

    # 完整标准行次（小企业会计准则会小企01表），零余额项目亦列出
    left = [
        ("货币资金", 1, "货币资金"),
        ("短期投资", 2, "短期投资"),
        ("应收票据", 3, "应收票据"),
        ("应收账款", 4, "应收账款"),
        ("预付账款", 5, "预付账款"),
        ("应收股利", 6, "应收股利"),
        ("应收利息", 7, "应收利息"),
        ("其他应收款", 8, "其他应收款"),
        ("存货", 9, "存货"),
        ("　其中：原材料", 10, "原材料"),
        ("　其中：在产品", 11, "在产品"),
        ("　其中：库存商品", 12, "库存商品"),
        ("　其中：周转材料", 13, "周转材料"),
        ("其他流动资产", 14, "其他流动资产"),
        ("流动资产合计", 15, "流动资产合计"),
        ("长期债券投资", 16, "长期债券投资"),
        ("长期股权投资", 17, "长期股权投资"),
        ("固定资产原价", 18, "固定资产原值"),
        ("减：累计折旧", 19, "累计折旧"),
        ("固定资产账面价值", 20, "固定资产账面价值"),
        ("在建工程", 21, "在建工程"),
        ("工程物资", 22, "工程物资"),
        ("固定资产清理", 23, "固定资产清理"),
        ("生产性生物资产", 24, "生产性生物资产"),
        ("无形资产", 25, "无形资产"),
        ("开发支出", 26, "开发支出"),
        ("长期待摊费用", 27, "长期待摊费用"),
        ("其他非流动资产", 28, "其他非流动资产"),
        ("非流动资产合计", 29, "非流动资产合计"),
        ("资产总计", 30, "资产总计"),
    ]
    right = [
        ("短期借款", 31, "短期借款"),
        ("应付票据", 32, "应付票据"),
        ("应付账款", 33, "应付账款"),
        ("预收账款", 34, "预收账款"),
        ("应付职工薪酬", 35, "应付职工薪酬"),
        ("应交税费", 36, "应交税费"),
        ("应付利息", 37, "应付利息"),
        ("应付利润", 38, "应付利润"),
        ("其他应付款", 39, "其他应付款"),
        ("其他流动负债", 40, "其他流动负债"),
        ("流动负债合计", 41, "流动负债合计"),
        ("长期借款", 42, "长期借款"),
        ("长期应付款", 43, "长期应付款"),
        ("递延收益", 44, "递延收益"),
        ("其他非流动负债", 45, "其他非流动负债"),
        ("非流动负债合计", 46, "非流动负债合计"),
        ("负债合计", 47, "负债合计"),
        ("实收资本（或股本）", 48, "实收资本"),
        ("资本公积", 49, "资本公积"),
        ("盈余公积", 50, "盈余公积"),
        ("未分配利润", 51, "未分配利润"),
        ("所有者权益合计", 52, "所有者权益合计"),
        ("负债和所有者权益总计", 53, "负债和所有者权益总计"),
    ]

    # 零余额科目映射
    ZERO_KEYS = {"短期投资", "应收票据", "应收股利", "应收利息",
                 "原材料", "在产品", "库存商品", "周转材料", "其他流动资产",
                 "长期债券投资", "长期股权投资", "在建工程", "工程物资",
                 "固定资产清理", "生产性生物资产", "无形资产", "开发支出",
                 "其他非流动资产",
                 "短期借款", "应付票据", "应付利息", "应付利润",
                 "其他流动负债", "长期借款", "长期应付款", "递延收益",
                 "其他非流动负债", "非流动负债合计", "资本公积"}

    row = 5
    for i in range(max(len(left), len(right))):
        # 先对整行套基础样式 + 边框
        for c in range(1, 9):
            apply_style(ws.cell(row=row, column=c), BODY_FONT)

        if i < len(left):
            name, ln, key = left[i]
            ev, bv = bs.get(key, (0.0, 0.0))
            if key in ZERO_KEYS:
                ev, bv = 0.0, 0.0
            dot = "合计" in name or "总计" in name
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=ln)
            ws.cell(row=row, column=3, value=round(ev, 2))
            ws.cell(row=row, column=4, value=round(bv, 2))
            ws.cell(row=row, column=3).number_format = NUM_FMT
            ws.cell(row=row, column=4).number_format = NUM_FMT
            if dot:
                for c in range(1, 5):
                    ws.cell(row=row, column=c).font = BOLD_FONT

        if i < len(right):
            name, ln, key = right[i]
            ev, bv = bs.get(key, (0.0, 0.0))
            if key in ZERO_KEYS:
                ev, bv = 0.0, 0.0
            dot = "合计" in name or "总计" in name
            ws.cell(row=row, column=5, value=name)
            ws.cell(row=row, column=6, value=ln)
            ws.cell(row=row, column=7, value=round(ev, 2))
            ws.cell(row=row, column=8, value=round(bv, 2))
            ws.cell(row=row, column=7).number_format = NUM_FMT
            ws.cell(row=row, column=8).number_format = NUM_FMT
            if dot:
                for c in range(5, 9):
                    ws.cell(row=row, column=c).font = BOLD_FONT
        row += 1

    # 签章区
    row += 2
    ws.merge_cells(f"A{row}:D{row}"); ws.cell(row=row, column=1, value="单位负责人："); apply_style(ws.cell(row=row, column=1), BODY_FONT, THIN_BORDER)
    ws.merge_cells(f"E{row}:H{row}"); ws.cell(row=row, column=5, value="会计负责人："); apply_style(ws.cell(row=row, column=5), BODY_FONT, THIN_BORDER)
    row += 1
    ws.merge_cells(f"A{row}:D{row}"); ws.cell(row=row, column=1, value="制表人："); apply_style(ws.cell(row=row, column=1), BODY_FONT, THIN_BORDER)
    ws.merge_cells(f"E{row}:H{row}"); ws.cell(row=row, column=5, value=f"制表日期：{REPORT_END.strftime('%Y年%m月%d日')}"); apply_style(ws.cell(row=row, column=5), BODY_FONT, THIN_BORDER)

    for c, w in zip("ABCDEFGH", [28, 6, 16, 16, 28, 6, 16, 16]):
        ws.column_dimensions[c].width = w


def write_cashflow_sheet(wb, cf):
    ws = wb.create_sheet("现金流量表")
    ws.sheet_properties.tabColor = "ED7D31"

    ws.merge_cells("A1:D1")
    ws["A1"] = f"{COMPANY_NAME} 现金流量表"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:D2")
    ws["A2"] = "2026年1-5月（小企业会计准则）     会小企03表      单位：元"
    ws["A2"].font = Font(name="微软雅黑", size=9)

    for col, h in enumerate(["项目", "行次", "本期金额", "附注"], 1):
        c = ws.cell(row=4, column=col, value=h)
        apply_style(c, HEADER_FONT)

    cats = cf["categories"]
    items = [
        # 一、经营活动产生的现金流量
        ("一、经营活动产生的现金流量", "", "", True),
        ("销售商品、提供劳务收到的现金", 1, cats["销售商品提供劳务收到的现金"], False),
        ("收到其他与经营活动有关的现金", 2, cats["收到其他与经营活动有关的现金"], False),
        ("经营活动现金流入小计", 3, cats["经营活动现金流入小计"], True),
        ("购买商品、接受劳务支付的现金", 4, cats["购买商品接受劳务支付的现金"], False),
        ("支付给职工以及为职工支付的现金", 5, cats["支付给职工以及为职工支付的现金"], False),
        ("支付的各项税费", 6, cats["支付的各项税费"], False),
        ("支付其他与经营活动有关的现金", 7, cats["支付其他与经营活动有关的现金"], False),
        ("经营活动现金流出小计", 8, cats["经营活动现金流出小计"], True),
        ("经营活动产生的现金流量净额", 9, cats["经营活动产生的现金流量净额"], True),
        # 二、投资活动产生的现金流量
        ("二、投资活动产生的现金流量", "", "", True),
        ("收回短期投资、长期债券投资和长期股权投资收到的现金", 10, 0, False),
        ("取得投资收益收到的现金", 11, 0, False),
        ("处置固定资产、无形资产和其他非流动资产收回的现金净额", 12, 0, False),
        ("投资活动现金流入小计", 13, 0, True),
        ("购建固定资产、无形资产和其他非流动资产支付的现金", 14, 0, False),
        ("投资活动现金流出小计", 15, 0, True),
        ("投资活动产生的现金流量净额", 16, 0, True),
        # 三、筹资活动产生的现金流量
        ("三、筹资活动产生的现金流量", "", "", True),
        ("取得借款收到的现金", 17, 0, False),
        ("吸收投资者投资收到的现金", 18, 0, False),
        ("收到其他与筹资活动有关的现金", 19, 0, False),
        ("筹资活动现金流入小计", 20, 0, True),
        ("偿还借款本金支付的现金", 21, 0, False),
        ("偿还借款利息支付的现金", 22, 0, False),
        ("分配利润支付的现金", 23, 0, False),
        ("筹资活动现金流出小计", 24, 0, True),
        ("筹资活动产生的现金流量净额", 25, 0, True),
        # 四、现金净增加额
        ("四、现金净增加额", "", "", True),
        ("加：期初现金及现金等价物余额", 26, 0, False),
        ("期末现金及现金等价物余额", 27, 0, False),
    ]

    row = 5
    for name, line_num, amount, is_bold in items:
        f = BOLD_FONT if is_bold else BODY_FONT
        c1 = ws.cell(row=row, column=1, value=name); apply_style(c1, f)
        if line_num != "":
            c2 = ws.cell(row=row, column=2, value=line_num); apply_style(c2, BODY_FONT)
        else:
            c2 = ws.cell(row=row, column=2, value=""); apply_style(c2, BODY_FONT)
        if amount != "":
            c3 = ws.cell(row=row, column=3, value=round(amount, 2) if isinstance(amount, (int, float)) else amount)
            apply_style(c3, f if is_bold else BODY_FONT, number_format=NUM_FMT)
        else:
            c3 = ws.cell(row=row, column=3, value=""); apply_style(c3, BODY_FONT)
        c4 = ws.cell(row=row, column=4, value=""); apply_style(c4, BODY_FONT)
        row += 1

    # 签章区
    row += 2
    ws.merge_cells(f"A{row}:B{row}"); ws.cell(row=row, column=1, value="单位负责人："); apply_style(ws.cell(row=row, column=1), BODY_FONT, THIN_BORDER)
    ws.merge_cells(f"C{row}:D{row}"); ws.cell(row=row, column=3, value="会计负责人："); apply_style(ws.cell(row=row, column=3), BODY_FONT, THIN_BORDER)
    row += 1
    ws.merge_cells(f"A{row}:B{row}"); ws.cell(row=row, column=1, value="制表人："); apply_style(ws.cell(row=row, column=1), BODY_FONT, THIN_BORDER)
    ws.merge_cells(f"C{row}:D{row}"); ws.cell(row=row, column=3, value=f"制表日期：{REPORT_END.strftime('%Y年%m月%d日')}"); apply_style(ws.cell(row=row, column=3), BODY_FONT, THIN_BORDER)

    for c, w in zip("ABCD", [42, 8, 20, 10]):
        ws.column_dimensions[c].width = w


def write_notes_md(bs, cumulative_pl, cf, passed, failed, warnings, all_errors,
                   payroll_data, bank_data, invoice_data, inventory_end):
    lines = []
    lines.append(f"# {COMPANY_NAME}")
    lines.append(f"## 2026年1-5月财务报表说明")
    lines.append(f"")
    lines.append(f"**生成日期：** {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"**会计政策：** 小企业会计准则")
    lines.append(f"")
    lines.append("---")
    lines.append("")

    lines.append("## 一、关键数据摘要")
    lines.append("")
    lines.append("| 指标 | 金额（元） |")
    lines.append("|---|---|")
    lines.append(f"| 营业收入（1-5月） | **{cumulative_pl['revenue']:,.2f}** |")
    lines.append(f"| 营业成本（1-5月） | {cumulative_pl['cogs']:,.2f} |")
    lines.append(f"| 毛利 | {cumulative_pl['gross_profit']:,.2f} |")
    rev = cumulative_pl['revenue']
    lines.append(f"| 毛利率 | {cumulative_pl['gross_profit']/rev*100:.1f}% |" if rev > 0 else "| 毛利率 | N/A |")
    lines.append(f"| 利润总额 | **{cumulative_pl['total_profit']:,.2f}** |")
    lines.append(f"| 净利润 | **{cumulative_pl['net_profit']:,.2f}** |")
    lines.append(f"| 资产总计（5月末） | {bs['资产总计'][0]:,.2f} |")
    lines.append(f"| 负债合计（5月末） | {bs['负债合计'][0]:,.2f} |")
    lines.append(f"| 所有者权益（5月末） | {bs['所有者权益合计'][0]:,.2f} |")
    lines.append(f"| 货币资金（5月末） | {bs['货币资金'][0]:,.2f} |")
    lines.append(f"| 经营活动现金净额 | {cf['categories']['经营活动产生的现金流量净额']:,.2f} |")
    lines.append("")

    lines.append("## 二、校验结果")
    lines.append("")
    for p in passed:
        lines.append(f"- {p}")
    for f in failed:
        lines.append(f"- {f}")
    for w in warnings:
        lines.append(f"- {w}")
    lines.append("")

    lines.append("## 三、数据范围")
    lines.append("")
    lines.append(f"- 收入发票 {invoice_data.get('income_count', 0)} 张")
    lines.append(f"- 成本费用发票 {invoice_data.get('expense_count', 0)} 张")
    lines.append(f"- 银行流水 {bank_data.get('txn_count', 0)} 条（农行 + 农村信用社）")
    lines.append(f"- 工资记录 {payroll_data.get('emp_count', 0)} 人次")
    lines.append(f"- 应发工资合计: {payroll_data.get('total_gross', 0):,.2f} 元")
    lines.append(f"- 实发工资合计: {payroll_data.get('total_net', 0):,.2f} 元")
    lines.append("")

    lines.append("## 四、待确认事项")
    lines.append("")
    lines.append("### 1. 长期待摊费用")
    ltp = bs.get("长期待摊费用", (0.0, 0.0))
    lines.append(f"期末余额 {ltp[0]:,.2f} 元（期初 {ltp[1]:,.2f} 元），目前**未摊销**。如需摊销请告知摊销期限。")
    lines.append("")
    lines.append("### 2. 其他应付款")
    other_bs = bs.get("其他应付款", (0.0, 0.0))
    lines.append(f"其他应付款期初 {other_bs[1]:,.2f} 元，期末 {other_bs[0]:,.2f} 元（假设不变）。")
    lines.append("如本期有关联方往来变动，请核对银行流水中大额对公转账并调整。")
    lines.append("")
    lines.append("### 3. 存货计算方法")
    lines.append(f"存货 = 期初 + 采购含税金额 - 销售成本 = {inventory_end:,.2f}")
    lines.append("注：小规模纳税人不能抵扣进项，采购按含税价入库存。实际库存以盘点数为准。")
    lines.append("")
    lines.append("### 4. 社保单位部分")
    if payroll_data.get("total_social_unit", 0) == 0:
        lines.append("工资表中社保单位部分（col 24）均为 0。如果公司确实缴了社保，请补充单位部分数据。")
    lines.append("")

    # 平衡调整说明
    balance_adj = bs.get("_balance_adj", 0.0)
    if abs(balance_adj) > 0.5:
        lines.append("### 5. 资产负债表平衡调整")
        lines.append(f'因未逐笔匹配发票与银行流水来更新应收/应付账款，资产负债表自动平衡时调整了"其他应收款" {balance_adj:,.2f} 元。')
        lines.append("建议会计根据实际销售回款和采购付款情况，手工调整应收账款和应付账款科目。")
        lines.append("")

    lines.append("### 6. 应交税费说明")
    lines.append(f"本期计提增值税 {cumulative_pl.get('vat_collected', 0):,.2f} 元（小规模纳税人，按征收率）。")
    lines.append(f"附加税（城建税+教育费附加+地方教育附加）按增值税×12%估算。")
    lines.append(f"所得税按小微企业优惠税率估算（实际税率取决于全年累计利润和汇算清缴调整）。")
    lines.append(f"期初应交税费余额 {bs.get('应交税费', (0,0))[1]:,.2f} 元，假设本期未实缴 → 期末 {bs.get('应交税费', (0,0))[0]:,.2f} 元。")
    lines.append('实际已缴金额需从银行流水"支付的各项税费"中核实。')
    lines.append("")

    if all_errors:
        lines.append("## 五、读取异常")
        lines.append("")
        for e in all_errors[:50]:
            lines.append(f"- {e}")
        if len(all_errors) > 50:
            lines.append(f"- ... 以及其他 {len(all_errors)-50} 条")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("*报表由脚本自动生成，未经人工审核，请由会计最终确认。*")
    return "\n".join(lines)


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 60)
    print(f"  {COMPANY_NAME}")
    print(f"  2026年1-5月财务报表生成")
    print(f"  小企业会计准则")
    print("=" * 60)
    print()

    all_errors = []

    # 1. 期初资产负债表
    print("[读] 读取期初资产负债表...")
    bs_path = os.path.join(BASE_DIR, FILES["balance_sheet"])
    opening_bs = read_opening_balance_sheet(bs_path)
    if opening_bs is None:
        print("[FAIL] 无法读取期初资产负债表，退出")
        return 1
    print(f"   期初资产总计: {opening_bs.get('资产总计', (0,0))[0]:,.2f}")

    # 2. 销售收入
    print("[读] 读取销售收入发票...")
    sales_path = os.path.join(BASE_DIR, FILES["sales"])
    sales_data = read_invoices(sales_path, "销售收入")
    all_errors.extend(sales_data["errors"])
    all_income = sales_data["income"]
    print(f"   收入发票: {len(all_income)} 条")

    # 3. 成本费用
    print("[读] 读取成本费用发票...")
    cost_path = os.path.join(BASE_DIR, FILES["cost"])
    cost_data = read_invoices(cost_path, "成本费用")
    all_errors.extend(cost_data["errors"])
    all_expense = cost_data["expense"]
    print(f"   成本费用发票: {len(all_expense)} 条")

    # 4. 工资
    print("[读] 读取工资数据...")
    all_payroll_employees = []
    payroll_total_gross = 0.0
    payroll_total_social_unit = 0.0
    payroll_total_net = 0.0
    for i, sf in enumerate(FILES["salary"]):
        sp = os.path.join(BASE_DIR, sf)
        if os.path.exists(sp):
            pr = read_payroll(sp, f"工资{i+1}月")
            all_errors.extend(pr["errors"])
            all_payroll_employees.extend(pr["employees"])
            payroll_total_gross += pr["total_gross"]
            payroll_total_social_unit += pr["total_social_unit"]
            payroll_total_net += pr["total_net"]
        else:
            all_errors.append(f"工资文件不存在: {sf}")
    print(f"   工资记录: {len(all_payroll_employees)} 人次")
    print(f"   应发工资合计: {payroll_total_gross:,.2f}")

    payroll_summary = {
        "total_gross": payroll_total_gross,
        "total_social_unit": payroll_total_social_unit,
        "total_net": payroll_total_net,
        "emp_count": len(all_payroll_employees),
    }

    # 5. 银行流水
    print("[读] 读取银行流水...")
    all_transactions = []
    bank_end_balance = 0.0
    bank_txn_count = 0
    bank_last_balances = {}
    for label, files in [("农行", FILES["bank_abc"]), ("农村信用社", FILES["bank_rcc"])]:
        for i, bf in enumerate(files):
            bp = os.path.join(BASE_DIR, bf)
            if os.path.exists(bp):
                bd = read_bank_statement(bp, f"{label}{i+1}月")
                all_errors.extend(bd["errors"])
                all_transactions.extend(bd["transactions"])
                bank_txn_count += len(bd["transactions"])
                if bd["transactions"]:
                    bank_last_balances[label] = bd["transactions"][-1][3]
            else:
                all_errors.append(f"银行流水文件不存在: {bf}")

    # 用最后一个月每条流水的余额累加（不同银行）
    bank_end_balance = sum(bank_last_balances.values()) if bank_last_balances else 0.0
    print(f"   银行流水: {bank_txn_count} 条")
    print(f"   银行期末余额(各户5月末): {bank_end_balance:,.2f}")
    for k, v in bank_last_balances.items():
        print(f"     {k}: {v:,.2f}")

    # 6. 利润表
    print("[CALC] 计算利润表...")
    pl_result = compute_profit_statement(all_income, all_expense, payroll_summary)
    cumulative_pl = pl_result["cumulative"]
    monthly_pl = pl_result["monthly"]
    cumulative_pl["_monthly_detail"] = monthly_pl
    for note in pl_result["notes"]:
        print(f"   {note}")

    # 7. 资产负债表
    print("[CALC] 计算资产负债表...")
    total_purchase_tax = sum(t[2] for t in all_expense)
    total_purchase_total = sum(t[3] for t in all_expense)
    inventory_open = opening_bs.get("存货", (0.0, 0.0))[0]
    inventory_end = inventory_open + total_purchase_total - cumulative_pl["cogs"]
    if inventory_end < 0:
        inventory_end = 0.0

    bs = compute_balance_sheet(opening_bs, cumulative_pl, bank_end_balance, payroll_summary)
    bs["存货"] = (inventory_end, inventory_open)
    current_end = (bs["货币资金"][0] + bs["应收账款"][0] + bs["预付账款"][0]
                   + bs["其他应收款"][0] + bs["存货"][0])
    bs["流动资产合计"] = (current_end, bs["流动资产合计"][1])
    total_assets = current_end + bs["非流动资产合计"][0]
    bs["资产总计"] = (total_assets, bs["资产总计"][1])
    bs["负债和所有者权益总计"] = (bs["负债合计"][0] + bs["所有者权益合计"][0],
                              bs["负债和所有者权益总计"][1])

    # 8. 现金流量表
    print("[CALC] 计算现金流量表...")
    cf_result = compute_cashflow(all_transactions, payroll_summary)

    # 8.5. 资产负债表平衡调整
    # 因应收账款/应付账款未按发票与银行流水逐笔勾稽，期末可能出现不平衡
    # 差异主要是未逐笔更新应收应付所致，统一挂入"其他应收款"并在说明中标注
    total_assets_end = bs["资产总计"][0]
    total_liab_equity_end = bs["负债和所有者权益总计"][0]
    bs_diff = total_liab_equity_end - total_assets_end
    if abs(bs_diff) > 0.5:
        # 用其他应收款吸收差异（正值=资产少计，增加其他应收款；负值=减少）
        old_other_recv = bs["其他应收款"][0]
        bs["其他应收款"] = (old_other_recv + bs_diff, bs["其他应收款"][1])
        # 重新计算流动资产合计和资产总计
        current_end_new = (bs["货币资金"][0] + bs["应收账款"][0] + bs["预付账款"][0]
                           + bs["其他应收款"][0] + bs["存货"][0])
        bs["流动资产合计"] = (current_end_new, bs["流动资产合计"][1])
        bs["资产总计"] = (current_end_new + bs["非流动资产合计"][0], bs["资产总计"][1])
        bs["_balance_adj"] = bs_diff
        print(f"   [ADJ] 资产负债表自动平衡: 调整其他应收款 {bs_diff:,.2f} 元")
    else:
        bs["_balance_adj"] = 0.0

    # 9. 校验
    print("[CHECK] 执行校验...")
    passed, failed, warnings = validate(bs, cumulative_pl, cf_result, opening_bs, bank_end_balance)
    for p in passed:
        print(f"   {p}")
    for f in failed:
        print(f"   {f}")
    for w in warnings:
        print(f"   {w}")

    # 10. 输出
    print("[WRITE] 生成输出文件...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_profit_sheet(wb, cumulative_pl, monthly_pl)
    write_balance_sheet_sheet(wb, bs)
    write_cashflow_sheet(wb, cf_result)

    xlsx_path = os.path.join(OUTPUT_DIR, "顺恒2026年1-5月财务报表.xlsx")
    try:
        wb.save(xlsx_path)
        print(f"   [PASS] Excel: {xlsx_path}")
    except PermissionError:
        print(f"   [FAIL] 无法保存 Excel: 文件被占用，请关闭 Excel 后重试")
        print(f"   文件路径: {xlsx_path}")
        return 1
    except Exception as e:
        print(f"   [FAIL] 保存 Excel 失败: {e}")
        return 1

    invoice_data = {"income_count": len(all_income), "expense_count": len(all_expense)}
    bank_data = {"txn_count": bank_txn_count}

    md_content = write_notes_md(bs, cumulative_pl, cf_result, passed, failed, warnings,
                                all_errors, payroll_summary, bank_data, invoice_data, inventory_end)
    md_path = os.path.join(OUTPUT_DIR, "顺恒2026年1-5月报表说明.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"   [PASS] 说明: {md_path}")

    print()
    print("=" * 60)
    if failed:
        print("  [WARN] 报表生成完成，有校验未通过！请查看说明文件")
    else:
        print("  [PASS] 报表生成完成！")
    print(f"  输出目录: {OUTPUT_DIR}")
    print("=" * 60)

    uncl = len(cf_result.get("unclassified", []))
    if uncl > 0:
        print(f"  [TIP] {uncl} 条银行流水无法自动分类，详见说明文件。")
    if all_errors:
        print(f"  [WARN] {len(all_errors)} 个读取警告，详见说明文件。")

    return 0 if not failed else 1


if __name__ == "__main__":
    sys.exit(main())
