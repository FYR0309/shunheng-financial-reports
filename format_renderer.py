# -*- coding: utf-8 -*-
"""Format renderer — produces formatted .xlsx files from template config + data.

Adapts formatting constants and logic from generate_formatted.py (lines 219-575).
Driven entirely by template JSON configs — no hardcoded company-specific data.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -------- Format constants (verified against Dec 2025 templates) --------

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
no_border = Border()

font_title = Font(name='宋体', size=14)
font_subtitle = Font(name='宋体', size=10)
font_header = Font(name='宋体', size=10)
font_item = Font(name='宋体', size=10)
font_section = Font(name='宋体', size=10)
font_number = Font(name='Arial', size=10)

align_title = Alignment(horizontal='center', vertical='center')
align_subtitle_right = Alignment(horizontal='right', vertical='center')
align_header = Alignment(horizontal='center', vertical='center')
align_item_left = Alignment(horizontal='left', vertical='center')
align_number = Alignment(horizontal='right', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')

num_fmt = '#,##0.00'


# -------- Renderer --------

class ReportRenderer:
    """Render one statement type to a formatted .xlsx workbook."""

    def __init__(self, template_config):
        self.cfg = template_config
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def render(self, company_name, period_str, data_map, output_path):
        """Render the full statement.

        Args:
            company_name: e.g. '来宾市顺恒废旧汽车回收有限公司'
            period_str: e.g. '2026年1-5月' or '2026年5月31日'
            data_map: dict mapping item_name -> {value_col_index: value}
                      e.g. {'一、营业收入': {0: 4080824.78, 1: 915170.87}}
            output_path: where to save the .xlsx file
        """
        cfg = self.cfg
        ncols = cfg['ncols']

        # -- Sheet title --
        self.ws.title = cfg['template_type']

        # Row 1: Title
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = self.ws.cell(row=1, column=1, value=cfg['template_type'])
        self._fmt(c, font_title, align_title, no_border)
        self.ws.row_dimensions[1].height = cfg.get('row_height_title', 21)

        # Row 2: Subtitle
        self.ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = self.ws.cell(row=2, column=1, value=cfg['subtitle'])
        self._fmt(c, font_subtitle, align_subtitle_right, no_border)
        self.ws.row_dimensions[2].height = cfg.get('row_height_data', 17.25)

        # Row 3: Company info
        self._write_company_row(3, company_name, period_str, ncols)

        # Row 4: Column headers
        self._write_col_headers(4, ncols)

        # Data rows
        self._write_data_rows(data_map, ncols)

        # Empty rows to reach total_rows
        data_end_row = cfg['data_start_row'] + len(cfg['items']) - 1
        for er in range(data_end_row + 1, cfg['total_rows'] + 1):
            self._write_empty_row(er, ncols)

        # Column widths
        if 'col_widths' in cfg:
            for ci, w in enumerate(cfg['col_widths'], 1):
                self.ws.column_dimensions[get_column_letter(ci)].width = w

        # Print settings
        self.ws.sheet_properties.pageSetUpPr = \
            openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        self.ws.page_setup.orientation = 'landscape'

        self.wb.save(output_path)
        return output_path

    def _write_company_row(self, row, company_name, period_str, ncols):
        cfg = self.cfg
        template_type = cfg['template_type']

        if template_type == '资产负债表':
            self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            c = self.ws.cell(row=row, column=1, value=f'编制单位：{company_name}')
            self._fmt(c, font_subtitle, align_item_left, no_border)

            self.ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
            c = self.ws.cell(row=row, column=5,
                             value=f'日期：{period_str}              单位：元')
            self._fmt(c, font_subtitle, align_item_left, no_border)
        else:
            texts = [f'编制单位：{company_name}', f'期间：{period_str}']
            unit_col = 4 if ncols >= 4 else 3

            # Write all cells BEFORE merging (otherwise MergedCell objects block writes)
            for ci in range(1, ncols + 1):
                if ci == 1:
                    txt = texts[0]
                elif ci == 2:
                    txt = texts[1]
                elif ci == unit_col:
                    txt = '单位：元'
                else:
                    txt = ''
                c = self.ws.cell(row=row, column=ci, value=txt)
                self._fmt(c, font_subtitle, align_item_left, no_border)

            # Now merge cols 2-3 for period info
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

        self.ws.row_dimensions[row].height = cfg.get('row_height_data', 17.25)

    def _write_col_headers(self, row, ncols):
        cfg = self.cfg

        if cfg['template_type'] == '资产负债表':
            headers = ['资产', '行次', '期末余额', '年初余额',
                       '负债和所有者权益', '行次', '期末余额', '年初余额']
        else:
            labels = cfg.get('value_col_labels', ['本年累计金额', '本月金额'])
            headers = ['项目', '行次'] + labels
            if ncols == 5:
                headers.append('')

        for ci, hdr in enumerate(headers, 1):
            if ci > ncols:
                break
            c = self.ws.cell(row=row, column=ci, value=hdr)
            self._fmt(c, font_header, align_header, thin_border)
        self.ws.row_dimensions[row].height = cfg.get('row_height_data', 17.25)

    def _write_data_rows(self, data_map, ncols):
        cfg = self.cfg

        for item in cfg['items']:
            row = item['row']
            is_section = item.get('section', False)

            if cfg['template_type'] == '资产负债表':
                self._write_bs_row(row, item, data_map, ncols, is_section)
            else:
                self._write_simple_row(row, item, data_map, ncols, is_section)

            self.ws.row_dimensions[row].height = cfg.get('row_height_data', 17.25)

    def _write_simple_row(self, row, item, data_map, ncols, is_section):
        """Write a PL or CF data row (single-sided)."""
        name = item.get('name', '')
        row_num_str = item.get('row_num', '')

        if not name and not row_num_str:
            self._write_empty_row(row, ncols)
            return

        # Col 0: item name
        c = self.ws.cell(row=row, column=1, value=name)
        self._fmt(c, font_section if is_section else font_item,
                  align_item_left, thin_border)

        # Col 1: row number
        c = self.ws.cell(row=row, column=2, value=row_num_str if row_num_str else None)
        self._fmt(c, font_number, align_center, thin_border)

        # Cols 2+: amount values
        row_data = data_map.get(name, {})
        for vi in range(len(self.cfg['columns']['value_cols'])):
            vcol = self.cfg['columns']['value_cols'][vi]
            val = row_data.get(vi)
            c = self.ws.cell(row=row, column=vcol + 1, value=val)
            self._fmt(c, font_number, align_number, thin_border, num_fmt)

    def _write_bs_row(self, row, item, data_map, ncols, is_section):
        """Write a BS data row (double-sided: assets left, liabilities right)."""
        left_name = item.get('left', '')
        left_rn = item.get('left_row_num', '')
        right_name = item.get('right', '')
        right_rn = item.get('right_row_num', '')

        # Col 0: asset name (col 1)
        c = self.ws.cell(row=row, column=1, value=left_name)
        self._fmt(c, font_section if is_section else font_item,
                  align_item_left, thin_border)

        # Col 1: row num (col 2)
        c = self.ws.cell(row=row, column=2, value=left_rn if left_rn else None)
        self._fmt(c, font_number, align_center, thin_border)

        # Cols 2-3: asset amounts (cols 3-4)
        left_data = data_map.get(left_name, {})
        c = self.ws.cell(row=row, column=3, value=left_data.get(0))
        self._fmt(c, font_number, align_number, thin_border, num_fmt)
        c = self.ws.cell(row=row, column=4, value=left_data.get(1))
        self._fmt(c, font_number, align_number, thin_border, num_fmt)

        # Col 4: liability name (col 5)
        c = self.ws.cell(row=row, column=5, value=right_name if right_name else None)
        self._fmt(c, font_section if is_section else font_item,
                  align_item_left, thin_border)

        # Col 5: row num (col 6)
        c = self.ws.cell(row=row, column=6, value=right_rn if right_rn else None)
        self._fmt(c, font_number, align_center, thin_border)

        # Cols 6-7: liability amounts (cols 7-8)
        right_data = data_map.get(right_name, {})
        c = self.ws.cell(row=row, column=7, value=right_data.get(0))
        self._fmt(c, font_number, align_number, thin_border, num_fmt)
        c = self.ws.cell(row=row, column=8, value=right_data.get(1))
        self._fmt(c, font_number, align_number, thin_border, num_fmt)

    def _write_empty_row(self, row, ncols):
        for ci in range(1, ncols + 1):
            c = self.ws.cell(row=row, column=ci, value=None)
            self._fmt(c, font_item, None, thin_border)
        self.ws.row_dimensions[row].height = 13.5

    @staticmethod
    def _fmt(cell, font=None, alignment=None, border=None, number_format=None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format
